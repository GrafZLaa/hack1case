import os
import json
import base64
import io
import re
import copy
import traceback
import logging
import time
from datetime import datetime, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
from typing import Any, Dict, List, Optional, Tuple

import requests
import fitz
import barcode
from barcode.writer import ImageWriter
import openpyxl
from PIL import Image, ImageEnhance, ImageFilter, ImageOps, ImageChops
import pytesseract
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file

load_dotenv()
log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

app = Flask(__name__)


TESSERACT_PATH = os.getenv("TESSERACT_PATH", r"C:\Program Files\Tesseract-OCR\tesseract.exe")
if os.path.exists(TESSERACT_PATH):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

OLLAMA_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434").rstrip("/")
DEFAULT_MODEL_CANDIDATES = "qwen2.5vl:72b,qwen2.5vl:32b,qwen2.5vl:7b,llama3.2-vision"
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen2.5vl:32b")
MODEL_CANDIDATES = [
    m.strip()
    for m in os.getenv("OLLAMA_MODEL_CANDIDATES", DEFAULT_MODEL_CANDIDATES).split(",")
    if m.strip()
]
MAX_LLM_IMAGES = max(1, int(os.getenv("MAX_LLM_IMAGES", "4")))
LLM_TEXT_LIMIT = max(2000, int(os.getenv("LLM_TEXT_LIMIT", "14000")))
LLM_RETRY_TEXT_LIMIT = max(1200, int(os.getenv("LLM_RETRY_TEXT_LIMIT", "7000")))
LLM_FALLBACK_TEXT_LIMIT = max(800, int(os.getenv("LLM_FALLBACK_TEXT_LIMIT", "3200")))
LLM_PRIMARY_MAX_IMAGES = max(1, int(os.getenv("LLM_PRIMARY_MAX_IMAGES", "1")))
LLM_IMAGE_MAX_SIDE = max(900, int(os.getenv("LLM_IMAGE_MAX_SIDE", "1600")))
LLM_REQUEST_TIMEOUT = max(10, int(os.getenv("LLM_REQUEST_TIMEOUT", "55")))
LLM_CONNECT_TIMEOUT = max(1, int(os.getenv("LLM_CONNECT_TIMEOUT", "5")))
LLM_MODEL_FAILOVER_TRIES = max(1, int(os.getenv("LLM_MODEL_FAILOVER_TRIES", "1")))
LLM_MAX_TOTAL_SEC = max(20, int(os.getenv("LLM_MAX_TOTAL_SEC", "90")))
ENABLE_LLM = os.getenv("ENABLE_LLM", "0").strip().lower() in {"1", "true", "yes"}

OCR_LANG = os.getenv("OCR_LANG", "rus+eng")
OCR_PSMS = [p.strip() for p in os.getenv("OCR_PSMS", "3,6,11").split(",") if p.strip()]
if not OCR_PSMS:
    OCR_PSMS = ["3", "6", "11"]
OCR_PSMS = [psm for psm in dict.fromkeys(OCR_PSMS) if re.fullmatch(r"\d{1,2}", psm)]
if not OCR_PSMS:
    OCR_PSMS = ["3", "6", "11"]
OCR_WORKERS = max(1, int(os.getenv("OCR_WORKERS", "4")))
OCR_IF_EMBEDDED_CHARS = max(0, int(os.getenv("OCR_IF_EMBEDDED_CHARS", "120")))
PDF_RENDER_SCALE = max(1.3, float(os.getenv("PDF_RENDER_SCALE", "2.0")))
OCR_MAX_VARIANTS = max(1, int(os.getenv("OCR_MAX_VARIANTS", "3")))
OCR_QUALITY_SHORTCIRCUIT = float(os.getenv("OCR_QUALITY_SHORTCIRCUIT", "320"))
OCR_TESSERACT_TIMEOUT_SEC = max(2, int(os.getenv("OCR_TESSERACT_TIMEOUT_SEC", "8")))
OCR_DATE_TESSERACT_TIMEOUT_SEC = max(1, int(os.getenv("OCR_DATE_TESSERACT_TIMEOUT_SEC", "3")))
OCR_RELEASE_SCAN_BUDGET_SEC = max(2.0, float(os.getenv("OCR_RELEASE_SCAN_BUDGET_SEC", "5")))
PREVIEW_RENDER_SCALE = max(1.0, float(os.getenv("PREVIEW_RENDER_SCALE", "2.2")))
OLLAMA_TAGS_TIMEOUT_SEC = max(1, int(os.getenv("OLLAMA_TAGS_TIMEOUT_SEC", "3")))
OLLAMA_MODEL_CACHE_TTL_SEC = max(5, int(os.getenv("OLLAMA_MODEL_CACHE_TTL_SEC", "60")))
OLLAMA_HEALTH_FAIL_TTL_SEC = max(5, int(os.getenv("OLLAMA_HEALTH_FAIL_TTL_SEC", "25")))
REGISTRY_STATE_FILE = os.getenv("REGISTRY_STATE_FILE", os.path.join("data", "registry_state.json"))
CONTROL_SAMPLES_FILE = os.getenv("CONTROL_SAMPLES_FILE", os.path.join("samples", "control_samples.json"))
EVAL_MAX_WORKERS = max(1, int(os.getenv("EVAL_MAX_WORKERS", "2")))
EVAL_FAST_DEFAULT = os.getenv("EVAL_FAST_DEFAULT", "1").strip().lower() in {"1", "true", "yes", "on"}
EVAL_USE_CACHE_DEFAULT = os.getenv("EVAL_USE_CACHE_DEFAULT", "1").strip().lower() in {"1", "true", "yes", "on"}
MAX_REGISTRY_RECORDS = max(10, int(os.getenv("MAX_REGISTRY_RECORDS", "2000")))
REGISTRY_B64_MAX = max(12000000, int(os.getenv("REGISTRY_B64_MAX", "12000000")))
REGISTRY_MAX_IMAGES = max(1000, int(os.getenv("REGISTRY_MAX_IMAGES", "1000")))
FEEDBACK_FILE = os.getenv("FEEDBACK_FILE", os.path.join("data", "feedback_log.jsonl"))
OCR_KEYWORDS = [
    "паспорт",
    "руководство",
    "наименование",
    "издел",
    "заводск",
    "номер",
    "дата",
    "производител",
    "контроллер",
    "модул",
    "сертификат",
    "гарант",
    "срок",
    "служб",
    "код",
    "заказ",
]

RU_MONTHS = {
    "ЯНВ": "01",
    "ФЕВ": "02",
    "МАР": "03",
    "АПР": "04",
    "МАЙ": "05",
    "ИЮН": "06",
    "ИЮЛ": "07",
    "АВГ": "08",
    "СЕН": "09",
    "ОКТ": "10",
    "НОЯ": "11",
    "ДЕК": "12",
}

CYR_TO_LAT_LOOKALIKE = str.maketrans({
    "А": "A",
    "В": "B",
    "Е": "E",
    "К": "K",
    "М": "M",
    "Н": "H",
    "О": "O",
    "Р": "P",
    "С": "C",
    "Т": "T",
    "У": "Y",
    "Х": "X",
})

SERIAL_CONFUSABLES = {
    "1": "I",
    "I": "1",
    "L": "1",
    "5": "S",
    "S": "5",
    "8": "B",
    "B": "8",
    "6": "G",
    "G": "6",
    "2": "Z",
    "Z": "2",
}

PROMPT_EXTRACT = """You extract equipment-passport data from OCR text and page images.
Rules:
1) Use only text present in provided OCR/page snippets.
2) Never invent manufacturer/model values.
3) Keep Cyrillic exactly as in source where possible.
4) If field is absent, return null (or [] for arrays).

Return strict JSON with fields:
- document_type: one of ["single_passport", "group_passport", "cabinet_list", "unknown"]
- naimenovanie
- kod_dokumenta
- proizvoditel
- adres
- kontakty
- garantia
- srok_sluzhby
- sertifikat
- kod_zakaza
- normativnye_dok (array)
- komplektnost (array)
- zavodskie_nomera (array)
- data_vypuska
- data_priemki
"""

REGISTRY_SCALAR_FIELDS = [
    "document_type",
    "tip_pasporta",
    "naimenovanie",
    "kod_dokumenta",
    "kod_zakaza",
    "data_vypuska",
    "data_priemki",
    "proizvoditel",
    "adres",
    "kontakty",
    "garantia",
    "srok_sluzhby",
    "sertifikat",
    "_fileName",
]
REGISTRY_LIST_FIELDS = [
    "zavodskie_nomera",
    "normativnye_dok",
    "komplektnost",
]

_ollama_model_cache = {
    "model": "",
    "until": 0.0,
}
_ollama_health_cache = {
    "alive": True,
    "until": 0.0,
    "last_error": "",
}
_eval_extract_cache: Dict[str, Dict[str, Any]] = {}
_eval_extract_cache_lock = Lock()


def improve_image_variants(img_bytes: bytes) -> List[Image.Image]:
    """Build a compact set of preprocessed variants to improve OCR hit rate."""
    variants: List[Image.Image] = []
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("L")
        img = ImageOps.autocontrast(img)

        max_side = max(img.size)
        if max_side < 1800:
            scale = 1.22
            img = img.resize((max(1, int(img.width * scale)), max(1, int(img.height * scale))), Image.Resampling.LANCZOS)
        elif max_side > 3000:
            scale = 3000.0 / float(max_side)
            img = img.resize((max(1, int(img.width * scale)), max(1, int(img.height * scale))), Image.Resampling.LANCZOS)

        denoised = img.filter(ImageFilter.MedianFilter(size=3))
        sharp = ImageEnhance.Sharpness(denoised).enhance(2.0)
        high_contrast = ImageEnhance.Contrast(sharp).enhance(1.6)
        bw = sharp.point(lambda x: 255 if x > 162 else 0, mode="1").convert("L")

        variants = [img, high_contrast, bw]
    except Exception:
        try:
            variants = [Image.open(io.BytesIO(img_bytes)).convert("L")]
        except Exception:
            variants = []
    return variants


def clean_json(text: str) -> Dict:
    if not text:
        return {}
    text = text.strip()
    text = re.sub(r"^```json\s*|\s*```$", "", text, flags=re.IGNORECASE | re.DOTALL)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                return {}
    return {}


def _is_placeholder_scalar(value: Any) -> bool:
    if value is None:
        return True
    s = re.sub(r"\s+", " ", str(value)).strip().lower()
    if not s:
        return True
    placeholders = {
        "-", "—", "_", "n/a", "na", "none", "null", "unknown", "неизвестно", "не указано", "нет",
        "?", "??", "...", "1", "l", "i", "|",
    }
    return s in placeholders


def _is_llm_payload_suspicious(parsed: Dict) -> bool:
    if not isinstance(parsed, dict) or not parsed:
        return True

    scalar_keys = [
        "naimenovanie",
        "kod_dokumenta",
        "proizvoditel",
        "adres",
        "kontakty",
        "garantia",
        "srok_sluzhby",
        "sertifikat",
        "kod_zakaza",
        "data_vypuska",
        "data_priemki",
    ]
    values = []
    for key in scalar_keys:
        raw = parsed.get(key)
        if raw is None:
            continue
        sval = re.sub(r"\s+", " ", str(raw)).strip()
        if not sval:
            continue
        values.append(sval)

    if not values:
        return False

    placeholder_count = sum(1 for v in values if _is_placeholder_scalar(v))
    if len(values) >= 3 and placeholder_count / float(len(values)) >= 0.6:
        return True

    lowered = [v.lower() for v in values]
    max_repeat = max(lowered.count(v) for v in set(lowered))
    repeated_value = max(set(lowered), key=lowered.count)
    if max_repeat >= 4 and repeated_value in {"1", "-", "—", "unknown", "null"}:
        return True

    one_char_count = sum(1 for v in values if len(v) <= 1)
    if len(values) >= 4 and one_char_count / float(len(values)) >= 0.5:
        return True

    return False


def normalize_serials(serials: Optional[List[str]]) -> List[str]:
    if not serials:
        return []
    seen = set()
    result = []
    for raw in serials:
        s = re.sub(r"\s+", "", str(raw or "").upper())
        s = s.strip(".,;:-")
        if re.fullmatch(r"[A-ZА-Я0-9\-]{3,24}", s):
            s = s.translate(CYR_TO_LAT_LOOKALIKE)
        if len(s) < 3:
            continue
        if s not in seen:
            seen.add(s)
            result.append(s)
    return result


def regex_fallbacks(text: str) -> Dict:
    data = {}
    doc_code = _pick_document_code(text)
    if doc_code:
        data["kod_dokumenta"] = doc_code

    data["zavodskie_nomera"] = _extract_serials(text)

    dates = []
    for m in re.finditer(r"\b([0-3]?\d[./][01]?\d[./](?:19|20)\d{2})\b", text):
        s = max(0, m.start() - 30)
        e = min(len(text), m.end() + 30)
        context = text[s:e].lower()
        if "версия" in context:
            continue
        if "дата" not in context:
            continue
        dates.append(m.group(1).replace("/", "."))

    text_date = _try_parse_textual_date(text)
    if text_date and text_date not in dates:
        dates.append(text_date)

    if dates:
        data.setdefault("data_vypuska", dates[0])
        if len(dates) > 1:
            data.setdefault("data_priemki", dates[1])
    return data


def ocr_text_quality(text: str, avg_conf: float = 0.0) -> float:
    if not text:
        return float("-inf")
    words = re.findall(r"[A-Za-zА-Яа-я0-9]{2,}", text)
    if not words:
        return float("-inf")

    lower = text.lower()
    keyword_hits = sum(1 for kw in OCR_KEYWORDS if kw in lower)
    good_chars = len(re.findall(r"[A-Za-zА-Яа-я0-9\s.,:;()\-/+№%\"'«»]", text))
    noisy_chars = max(0, len(text) - good_chars)
    noise_ratio = noisy_chars / max(1, len(text))
    unique_words = len(set(w.lower() for w in words))
    unique_ratio = unique_words / max(1, len(words))
    very_short_words = sum(1 for w in words if len(w) <= 2)
    short_ratio = very_short_words / max(1, len(words))
    avg_word_len = sum(len(w) for w in words) / max(1, len(words))

    score = (
        avg_conf * 2.2
        + keyword_hits * 42
        + unique_ratio * 90
        + avg_word_len * 8
        - short_ratio * 55
        - noise_ratio * 45
    )
    if keyword_hits == 0:
        score -= 85
    return score


def ocr_page_text(page_img: bytes) -> str:
    best = ""
    best_score = float("-inf")
    variants = improve_image_variants(page_img)[:OCR_MAX_VARIANTS]
    for image in variants:
        for psm in OCR_PSMS:
            try:
                txt = pytesseract.image_to_string(
                    image,
                    lang=OCR_LANG,
                    config=f"--oem 1 --psm {psm}",
                    timeout=OCR_TESSERACT_TIMEOUT_SEC,
                )
                score = ocr_text_quality(txt)
                if score > best_score:
                    best = txt
                    best_score = score
                if best_score >= OCR_QUALITY_SHORTCIRCUIT:
                    break
            except Exception:
                continue
        if best_score >= OCR_QUALITY_SHORTCIRCUIT:
            break
    return best


def pdf_to_images_and_text(pdf_bytes: bytes) -> Tuple[List[bytes], List[str]]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    raw_text_blocks = []
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(PDF_RENDER_SCALE, PDF_RENDER_SCALE))
        pages.append(pix.tobytes("png"))
        raw_text_blocks.append(page.get_text("text") or "")
    return pages, raw_text_blocks


def get_available_ollama_models(timeout_sec: int = OLLAMA_TAGS_TIMEOUT_SEC) -> List[str]:
    resp = requests.get(f"{OLLAMA_URL}/api/tags", timeout=timeout_sec)
    resp.raise_for_status()
    _ollama_health_cache["alive"] = True
    _ollama_health_cache["last_error"] = ""
    _ollama_health_cache["until"] = time.monotonic() + OLLAMA_MODEL_CACHE_TTL_SEC
    models = resp.json().get("models", [])
    return [m.get("name", "") for m in models if m.get("name")]


def is_ollama_reachable() -> bool:
    now = time.monotonic()
    cached_until = float(_ollama_health_cache.get("until") or 0.0)
    if now < cached_until:
        return bool(_ollama_health_cache.get("alive"))

    try:
        resp = requests.get(f"{OLLAMA_URL}/api/tags", timeout=min(2, OLLAMA_TAGS_TIMEOUT_SEC))
        resp.raise_for_status()
        _ollama_health_cache["alive"] = True
        _ollama_health_cache["last_error"] = ""
        _ollama_health_cache["until"] = now + OLLAMA_MODEL_CACHE_TTL_SEC
        return True
    except Exception as e:
        _ollama_health_cache["alive"] = False
        _ollama_health_cache["last_error"] = str(e)
        _ollama_health_cache["until"] = now + OLLAMA_HEALTH_FAIL_TTL_SEC
        return False


def choose_active_ollama_model() -> str:
    now = time.monotonic()
    cached_model = _ollama_model_cache.get("model") or ""
    cached_until = float(_ollama_model_cache.get("until") or 0.0)
    if cached_model and now < cached_until:
        return cached_model

    try:
        available = get_available_ollama_models()
        if not available:
            _ollama_model_cache["model"] = OLLAMA_MODEL
            _ollama_model_cache["until"] = now + OLLAMA_MODEL_CACHE_TTL_SEC
            return OLLAMA_MODEL
        available_set = {m.lower(): m for m in available}
        ranked = [OLLAMA_MODEL] + MODEL_CANDIDATES
        for candidate in ranked:
            c = candidate.lower()
            if c in available_set:
                picked = available_set[c]
                _ollama_model_cache["model"] = picked
                _ollama_model_cache["until"] = now + OLLAMA_MODEL_CACHE_TTL_SEC
                return picked
            if ":" not in c:
                matched = next((available_set[a] for a in available_set if a.startswith(c + ":")), None)
                if matched:
                    _ollama_model_cache["model"] = matched
                    _ollama_model_cache["until"] = now + OLLAMA_MODEL_CACHE_TTL_SEC
                    return matched
        picked = available[0]
        _ollama_model_cache["model"] = picked
        _ollama_model_cache["until"] = now + OLLAMA_MODEL_CACHE_TTL_SEC
        return picked
    except Exception as e:
        _ollama_health_cache["alive"] = False
        _ollama_health_cache["last_error"] = str(e)
        _ollama_health_cache["until"] = time.monotonic() + OLLAMA_HEALTH_FAIL_TTL_SEC
        log.info("Failed to resolve Ollama model list: %s", e)
        _ollama_model_cache["model"] = OLLAMA_MODEL
        _ollama_model_cache["until"] = now + min(10, OLLAMA_MODEL_CACHE_TTL_SEC)
        return OLLAMA_MODEL


def resolve_llm_backend() -> Tuple[str, str]:
    if not ENABLE_LLM:
        return "disabled", "disabled"
    if not is_ollama_reachable():
        return "disabled", "disabled"
    return "ollama", choose_active_ollama_model()


def _response_error_text(resp: requests.Response) -> str:
    try:
        payload = resp.json()
        if isinstance(payload, dict):
            if payload.get("error"):
                return str(payload["error"])
            return json.dumps(payload, ensure_ascii=False)[:600]
    except Exception:
        pass
    return (resp.text or "")[:600]


def prepare_llm_images_b64(pages: List[bytes]) -> List[str]:
    encoded = []
    for page in pages[:MAX_LLM_IMAGES]:
        try:
            img = Image.open(io.BytesIO(page))
            max_side = max(img.size)
            if max_side > LLM_IMAGE_MAX_SIDE:
                scale = LLM_IMAGE_MAX_SIDE / float(max_side)
                new_size = (
                    max(1, int(img.width * scale)),
                    max(1, int(img.height * scale)),
                )
                img = img.resize(new_size, Image.Resampling.LANCZOS)
            out = io.BytesIO()
            img.save(out, format="PNG", optimize=True)
            encoded.append(base64.b64encode(out.getvalue()).decode())
        except Exception:
            encoded.append(base64.b64encode(page).decode())
    return encoded


def _prioritize_ocr_for_llm(ocr_text: str, limit: int) -> str:
    if not ocr_text:
        return ""
    lines = _clean_lines(ocr_text)
    if not lines:
        return ""

    priority_patterns = [
        r"\b(паспорт|руководств|наименован|издел|заводск\w*\s+номер|код\s+заказ|сертификат|гарант\w*|срок\s+служб\w*|дата)\b",
        r"\b(контроллер|модул|блок|шкаф|розетк|панел|мфк|tcc|tecon)\b",
        r"[А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?[А-ЯA-Z]{0,3}",
    ]

    selected: List[str] = []
    seen = set()
    for ln in lines[:80]:
        if ln not in seen:
            seen.add(ln)
            selected.append(ln)

    for ln in lines:
        if ln in seen:
            continue
        if any(re.search(pat, ln, flags=re.IGNORECASE) for pat in priority_patterns):
            seen.add(ln)
            selected.append(ln)
        if len(selected) >= 260:
            break

    text = "\n".join(selected)
    return text[:limit]


def _candidate_ollama_models(primary_model: str) -> List[str]:
    ranked = [primary_model] + MODEL_CANDIDATES
    candidates: List[str] = []
    seen = set()
    for m in ranked:
        key = (m or "").strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        candidates.append(m.strip())

    try:
        available = get_available_ollama_models(timeout_sec=min(OLLAMA_TAGS_TIMEOUT_SEC, 2))
    except Exception:
        available = []

    if not available:
        return candidates[:LLM_MODEL_FAILOVER_TRIES]

    available_map = {m.lower(): m for m in available}
    resolved: List[str] = []
    seen_resolved = set()
    for cand in candidates:
        c = cand.lower()
        picked = ""
        if c in available_map:
            picked = available_map[c]
        elif ":" not in c:
            picked = next((available_map[a] for a in available_map if a.startswith(c + ":")), "")
        if not picked:
            continue
        pkey = picked.lower()
        if pkey in seen_resolved:
            continue
        seen_resolved.add(pkey)
        resolved.append(picked)
        if len(resolved) >= LLM_MODEL_FAILOVER_TRIES:
            break

    return resolved or candidates[:LLM_MODEL_FAILOVER_TRIES]


def _should_skip_llm_for_file(source_name: str, raw_text: str) -> bool:
    name = (source_name or "").lower().replace("ё", "е")
    text = (raw_text or "").lower().replace("ё", "е")
    skip_markers = [
        "экранная форма",
        "задача",
        "логистика",
        "кейс",
        "служебная записка",
        "разбор паспорта",
        "критерии оценивания",
        "контекст задачи",
        "исходной таблицы",
    ]
    if any(marker in name for marker in skip_markers):
        return True
    if any(marker in text for marker in skip_markers):
        return True
    return False


def _has_strong_heuristic_result(data: Dict) -> bool:
    if not data:
        return False
    quality = int(data.get("quality_score") or 0)
    name_ok = bool(data.get("naimenovanie"))
    code_or_serial_ok = bool(data.get("kod_dokumenta")) or bool(data.get("zavodskie_nomera"))
    return quality >= 60 and name_ok and code_or_serial_ok


def _build_ollama_attempts(page_images_b64: List[str], ocr_text: str) -> List[Dict[str, Any]]:
    primary_images = page_images_b64[:min(len(page_images_b64), LLM_PRIMARY_MAX_IMAGES)]
    text_primary = _prioritize_ocr_for_llm(ocr_text, LLM_TEXT_LIMIT)
    text_retry = _prioritize_ocr_for_llm(ocr_text, LLM_RETRY_TEXT_LIMIT)
    text_fallback = _prioritize_ocr_for_llm(ocr_text, LLM_FALLBACK_TEXT_LIMIT)
    timeout_primary = max(20, LLM_REQUEST_TIMEOUT)
    timeout_retry = max(14, min(int(LLM_REQUEST_TIMEOUT * 0.55), 24))
    timeout_fallback = max(10, min(int(LLM_REQUEST_TIMEOUT * 0.45), 18))

    attempts = [
        {
            "images": primary_images,
            "ocr_text": text_primary,
            "format_json": True,
            "read_timeout": timeout_primary,
        },
        {
            "images": primary_images,
            "ocr_text": text_retry,
            "format_json": True,
            "read_timeout": timeout_retry,
        },
        {
            "images": primary_images,
            "ocr_text": text_fallback,
            "format_json": True,
            "read_timeout": timeout_fallback,
        },
    ]

    filtered = []
    for a in attempts:
        if not a["images"] and not a["ocr_text"]:
            continue
        filtered.append(a)
    return filtered


def run_ollama_extraction(
    page_images_b64: List[str],
    ocr_text: str,
    model_name: str,
    deadline: Optional[float] = None,
) -> Dict:
    attempts = _build_ollama_attempts(page_images_b64, ocr_text)
    last_error = ""
    for attempt in attempts:
        if deadline is not None and time.monotonic() >= deadline:
            break

        read_timeout = attempt["read_timeout"]
        if deadline is not None:
            remaining = int(deadline - time.monotonic())
            if remaining <= 3:
                break
            read_timeout = min(read_timeout, max(4, remaining))

        prompt = f"{PROMPT_EXTRACT}\n\nOCR_TEXT:\n{attempt['ocr_text']}"
        payload = {
            "model": model_name,
            "prompt": prompt,
            "images": attempt["images"],
            "stream": False,
            "options": {"temperature": 0.0, "num_predict": 420},
        }
        if attempt["format_json"]:
            payload["format"] = "json"

        resp = requests.post(
            f"{OLLAMA_URL}/api/generate",
            json=payload,
            timeout=(LLM_CONNECT_TIMEOUT, read_timeout),
        )
        if resp.ok:
            raw = resp.json().get("response", "")
            parsed = clean_json(raw)
            if parsed and not _is_llm_payload_suspicious(parsed):
                return parsed
            if parsed:
                last_error = "suspicious low-quality JSON from model response"
            else:
                last_error = "empty/invalid JSON from model response"
            continue

        detail = _response_error_text(resp)
        last_error = f"HTTP {resp.status_code}: {detail}"

    raise requests.RequestException(f"Ollama extraction failed: {last_error}")


def run_llm_extraction(page_images_b64: List[str], ocr_text: str, provider: str, model_name: str) -> Dict:
    errors: List[str] = []
    deadline = time.monotonic() + LLM_MAX_TOTAL_SEC
    for model in _candidate_ollama_models(model_name):
        if time.monotonic() >= deadline:
            break
        try:
            return run_ollama_extraction(page_images_b64, ocr_text, model, deadline=deadline)
        except requests.RequestException as e:
            errors.append(f"{model}: {e}")
            continue
    raise requests.RequestException(" | ".join(errors) if errors else "Ollama extraction failed")


def collect_ocr_texts(pages: List[bytes], embedded_text_blocks: List[str], enabled: bool = True) -> List[str]:
    if not enabled:
        return ["" for _ in pages]

    results = ["" for _ in pages]
    indexed_tasks = []
    workers = min(OCR_WORKERS, max(1, len(pages)))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for idx, page in enumerate(pages):
            embedded = (embedded_text_blocks[idx] if idx < len(embedded_text_blocks) else "").strip()
            if len(embedded) >= OCR_IF_EMBEDDED_CHARS:
                continue
            indexed_tasks.append((idx, pool.submit(ocr_page_text, page)))

        for idx, future in indexed_tasks:
            try:
                results[idx] = future.result()
            except Exception:
                results[idx] = ""

    return results


def _clean_lines(text: str) -> List[str]:
    lines = []
    for raw in (text or "").splitlines():
        line = re.sub(r"\s{2,}", " ", raw.replace("\t", " ")).strip()
        if line:
            lines.append(line)
    return lines


def _extract_value_after_label(lines: List[str], label_patterns: List[str], lookahead: int = 3) -> str:
    for idx, line in enumerate(lines):
        for pat in label_patterns:
            m = re.search(pat, line, flags=re.IGNORECASE)
            if not m:
                continue
            tail = line[m.end():].strip(" :;,-—№")
            if tail and len(tail) >= 4 and len(re.findall(r"[A-Za-zА-Яа-я]", tail)) >= 2:
                return tail
            for j in range(idx + 1, min(len(lines), idx + 1 + lookahead)):
                candidate = lines[j].strip(" :;,-—")
                if not candidate:
                    continue
                if any(re.search(lp, candidate, flags=re.IGNORECASE) for lp in label_patterns):
                    continue
                if len(candidate) < 4:
                    continue
                if len(re.findall(r"[A-Za-zА-Яа-я]", candidate)) < 2:
                    continue
                return candidate
    return ""


def _pick_document_code(text: str) -> str:
    suffix_map = {
        "PS": "ПС",
        "ПС": "ПС",
        "RE": "РЭ",
        "РЭ": "РЭ",
        "TU": "ТУ",
        "ТУ": "ТУ",
    }
    separated_suffix = re.search(
        r"\b([А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?)\s*(ПС|PS|РЭ|RE|ТУ|TU)\b",
        text,
        flags=re.IGNORECASE,
    )
    if separated_suffix:
        base = separated_suffix.group(1).strip()
        raw_suffix = separated_suffix.group(2).strip().upper().replace("Ё", "Е")
        suffix = suffix_map.get(raw_suffix, raw_suffix)
        return f"{base} {suffix}"

    patterns = [
        r"\b[А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?[А-ЯA-Z]{0,3}\b",
        r"\b[А-ЯA-Z]{2,6}\.\d{6}\.\d{3,4}\b",
        r"\b\d{4}\.\d{6}\.\d{4}\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return ""

def _extract_normative_docs(text: str) -> List[str]:
    found = []
    patterns = [
        r"\bГОСТ\s*[A-ZА-Я0-9.\-–/ ]{3,40}",
        r"\bТР\s*ТС\s*\d{3}/\d{4}\b",
        r"\bТР\s*ЕАЭС\s*\d{3}/\d{4}\b",
        r"\b[А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,4}\s*ТУ\b",
        r"\b(?:ОКПД|OKPD)\s*2[:\s]*\d{2}(?:\.\d{2,3}){1,4}\b",
        r"\b(?:ОКВЭД|OKVED)\s*2[:\s]*\d{2}(?:\.\d{1,2}){0,3}\b",
        r"\b(?:ТН\s*ВЭД|TN[\s-]*VED)(?:\s*(?:ЕАЭС|EAEU))?[:\s]*\d{6,10}\b",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            value = re.sub(r"\s{2,}", " ", m.group(0)).strip(" ,.;")
            if len(value) < 5:
                continue
            if value not in found:
                found.append(value)
    return found[:10]


def _is_probable_serial(token: str) -> bool:
    t = token.strip(" .,;:()[]{}").upper()
    if not t or not re.search(r"\d", t):
        return False
    if re.match(r"^(ТУ|TY|ТY)[A-ZА-Я0-9\-]{4,}$", t):
        return False
    if re.fullmatch(r"[A-ZА-Я]{2,8}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?[A-ZА-Я]{0,3}", t):
        return False
    if len(t) < 4 or len(t) > 20:
        return False
    if re.fullmatch(r"\d{1,5}", t):
        return False
    if re.fullmatch(r"\d{6,7}", t):
        return False
    if re.fullmatch(r"(19|20)\d{2}", t):
        return False
    if re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{2,4}", t):
        return False
    if re.fullmatch(r"\d{8,15}", t):
        return True
    if re.fullmatch(r"[A-ZА-Я]{2,4}\d{3,4}", t):
        return False
    if re.search(r"\d+[XХ]\d+[XХ]\d+", t):
        return False
    if not re.fullmatch(r"[A-ZА-Я0-9\-]{5,20}", t):
        return False
    digit_count = len(re.findall(r"\d", t))
    letter_count = len(re.findall(r"[A-ZА-Я]", t))
    if len(t) < 7 or digit_count < 3 or letter_count < 1:
        return False
    return bool(re.search(r"\d{4,}", t))


def _is_table_serial_candidate(token: str) -> bool:
    t = _normalize_serial_token(token)
    if not t:
        return False
    if re.fullmatch(r"\d{6,12}", t):
        return True
    if re.fullmatch(r"[A-Z]\d[A-Z]\d{3,8}", t):
        return True
    if not re.fullmatch(r"[A-Z0-9\-]{5,20}", t):
        return False
    digit_count = len(re.findall(r"\d", t))
    letter_count = len(re.findall(r"[A-Z]", t))
    if digit_count < 4:
        return False
    return letter_count <= digit_count


def _normalize_serial_token(token: str) -> str:
    t = re.sub(r"\s+", "", str(token or "").upper()).strip(" .,;:()[]{}")
    return t.translate(CYR_TO_LAT_LOOKALIKE)


def _serial_plausibility_score(token: str) -> int:
    t = _normalize_serial_token(token)
    score = 0
    if re.match(r"^[A-Z][0-9]", t):
        score += 4
    if re.match(r"^[A-Z][0-9][A-Z][0-9]{3,}$", t):
        score += 5
    if re.search(r"\d{4,}", t):
        score += 3
    if re.search(r"[A-Z]", t):
        score += 2
    if t and t[0].isdigit() and re.search(r"[A-Z]", t):
        score -= 3
    if t.count("-") > 1:
        score -= 1
    return score


def _serial_candidates(token: str) -> List[str]:
    base = _normalize_serial_token(token)
    if not re.search(r"[A-Z]", base):
        return [base]
    candidates = {base}
    for pos, ch in enumerate(base[:4]):
        alt = SERIAL_CONFUSABLES.get(ch)
        if not alt:
            continue
        candidates.add(base[:pos] + alt + base[pos + 1 :])
    return [c for c in candidates if c]


def _canonicalize_serial(token: str) -> str:
    candidates = []
    for cand in _serial_candidates(token):
        if _is_probable_serial(cand):
            candidates.append(cand)
    if not candidates:
        return _normalize_serial_token(token)
    return max(candidates, key=_serial_plausibility_score)


def _extract_serials(text: str) -> List[str]:
    serials: List[str] = []
    raw_upper = (text or "").upper()
    uppercase_text = re.sub(r"\s+", " ", raw_upper)
    uppercase_lines = [re.sub(r"\s+", " ", line).strip() for line in raw_upper.splitlines() if line.strip()]

    labeled_pattern = r"ЗАВОД\w*\s*НОМЕР\s*[:;№\-–—]*\s*([A-ZА-Я0-9\-]{4,20})"
    labeled_matches = list(re.finditer(labeled_pattern, uppercase_text, flags=re.IGNORECASE))
    for m in labeled_matches:
        token = _canonicalize_serial(m.group(1).strip())
        if _is_probable_serial(token):
            serials.append(token)

    has_serial_labels = bool(
        re.search(r"ЗАВОД\w*\s*НОМЕР|СЕРИЙН\w*\s*НОМЕР|ЗАВ\.\s*№", uppercase_text, flags=re.IGNORECASE)
    )
    table_mode = (
        len(labeled_matches) > 1
        or bool(re.search(r"№\s*П/?П|ПЕРЕЧЕНЬ\s+ДОКУМЕНТАЦИИ|СТРАНИЦ\s*/\s*ЛИСТОВ", uppercase_text, flags=re.IGNORECASE))
    )
    if table_mode and uppercase_lines:
        in_table = False
        rows_after_header = 0
        stop_line_re = re.compile(
            r"\b(ДАТА\s+УПАК|ОТВЕТСТВЕН|ДАТА\s+ПРИЕМ|ТЕХНИЧЕСКИЙ\s+КОНТРОЛЬ|СВЕДЕНИЯ\s+О\s+ИЗГОТОВИТЕЛЕ)\b",
            flags=re.IGNORECASE,
        )
        for line in uppercase_lines:
            if re.search(r"ЗАВОД\w*\s*НОМЕР", line, flags=re.IGNORECASE):
                in_table = True
                rows_after_header = 0
                continue
            if not in_table:
                continue
            if stop_line_re.search(line):
                in_table = False
                continue
            rows_after_header += 1
            if rows_after_header > 180:
                in_table = False
                continue
            for token in re.findall(r"\b[A-ZА-Я0-9\-]{5,20}\b", line):
                corrected = _canonicalize_serial(token)
                if re.fullmatch(r"\d{7,12}", corrected):
                    serials.append(corrected)
                    continue
                if _is_probable_serial(corrected) and _is_table_serial_candidate(corrected):
                    serials.append(corrected)

    numeric_tokens = re.findall(r"\b\d{7}\b", uppercase_text)
    if numeric_tokens:
        counts: Dict[str, int] = {}
        for t in numeric_tokens:
            counts[t] = counts.get(t, 0) + 1
        repeated = [
            token for token, cnt in counts.items()
            if cnt >= 2 and token.startswith("3")
        ]
        if len(repeated) >= 6:
            serials.extend(sorted(repeated))

    if not serials and has_serial_labels:
        for m in re.finditer(r"(ЗАВОД\w*\s*НОМЕР|СЕРИЙН\w*\s*НОМЕР|ЗАВ\.\s*№)", uppercase_text, flags=re.IGNORECASE):
            window = uppercase_text[m.end() : m.end() + 220]
            for token in re.findall(r"\b[A-ZА-Я0-9\-]{5,20}\b", window):
                corrected = _canonicalize_serial(token)
                if _is_probable_serial(corrected):
                    serials.append(corrected)

    return normalize_serials(serials)


def _extract_labeled_factory_serials(text: str) -> List[str]:
    if not text:
        return []
    upper = re.sub(r"\s+", " ", text.upper())
    pattern = re.compile(
        r"(?:ЗАВОД\w*\s*НОМЕР|СЕРИЙН\w*\s*НОМЕР|ЗАВ\.\s*№)\s*[:;№\-–—]*\s*([A-ZА-Я0-9\-]{4,20})",
        flags=re.IGNORECASE,
    )
    found: List[str] = []
    for m in pattern.finditer(upper):
        token = _canonicalize_serial(m.group(1).strip())
        if _is_probable_serial(token):
            found.append(token)
    return normalize_serials(found)


def _is_date_like(value: str) -> bool:
    if not value:
        return False
    if re.search(r"\b[0-3]?\d[./][01]?\d[./](?:19|20)\d{2}\b", value):
        return True
    if re.search(r"\b[0-3]?\d[./][01]?\d[./]\d{2}\b", value):
        return True
    if re.search(r"\b[0-3]?\d\s+[А-Яа-я]{3,}\s+(?:19|20)\d{2}\b", value, flags=re.IGNORECASE):
        return True
    return False


def _is_duration_like(value: str) -> bool:
    if not value:
        return False
    s = re.sub(r"\s+", " ", str(value).lower().replace("ё", "е")).strip(" .,:;")
    if len(s) < 2 or len(s) > 60:
        return False
    if _is_date_like(s):
        return True
    units = r"(год(?:а|ов)?|лет|месяц(?:а|ев)?|сут(?:ки|ок)?|дн(?:ей|я)?|день|недел(?:я|и|ь)|час(?:ов|а)?)"
    num_words = r"(один|одна|два|две|три|четыре|пять|шесть|семь|восемь|девять|десять|двенадцать|пятнадцать|двадцать|тридцать|сорок|пятьдесят)"
    has_amount = bool(re.search(r"\d", s) or re.search(num_words, s))
    return bool(has_amount and re.search(units, s))


def _normalize_numeric_date(day: str, month: str, year: str) -> str:
    d = max(1, min(31, int(day)))
    m = max(1, min(12, int(month)))
    y = int(year)
    if y < 100:
        y = 2000 + y if y <= 39 else 1900 + y
    return f"{d:02d}.{m:02d}.{y:04d}"


def _try_parse_textual_date(value: str) -> str:
    s = (value or "").upper().replace("Ё", "Е")
    s = s.translate({
        ord("A"): "А",
        ord("B"): "В",
        ord("C"): "С",
        ord("E"): "Е",
        ord("H"): "Н",
        ord("K"): "К",
        ord("M"): "М",
        ord("O"): "О",
        ord("P"): "Р",
        ord("T"): "Т",
        ord("X"): "Х",
        ord("Y"): "У",
    })
    m = re.search(r"\b([0-3]?\d)\s*[-./ ]?\s*([А-Я]{3,8})\.?\s*((?:19|20)?\d{2,4})\b", s)
    if not m:
        return ""
    day, month_word, year = m.group(1), m.group(2), m.group(3)
    month_key = next((k for k in RU_MONTHS if month_word.startswith(k)), "")
    if not month_key:
        return ""
    if len(year) == 4 and not year.startswith(("19", "20")):
        return ""
    if len(year) not in (2, 4):
        return ""
    return _normalize_numeric_date(day, RU_MONTHS[month_key], year)


def _extract_dates_from_chunk(chunk: str) -> List[str]:
    found = []
    for m in re.finditer(r"\b([0-3]?\d)[./]([01]?\d)[./]((?:19|20)?\d{2,4})\b", chunk):
        day, month, year = m.group(1), m.group(2), m.group(3)
        if len(year) == 4 and not year.startswith(("19", "20")):
            continue
        if len(year) not in (2, 4):
            continue
        try:
            found.append(_normalize_numeric_date(day, month, year))
        except Exception:
            continue

    textual = _try_parse_textual_date(chunk)
    if textual:
        found.append(textual)
    return list(dict.fromkeys(found))


def _extract_version_dates(text: str) -> set:
    result = set()
    for m in re.finditer(r"версия[^\n]{0,50}", text, flags=re.IGNORECASE):
        for d in _extract_dates_from_chunk(m.group(0)):
            result.add(d)
    return result


def _find_date_near_label(lines: List[str], label_patterns: List[str], window: int = 6) -> str:
    for idx, line in enumerate(lines):
        if not any(re.search(pat, line, flags=re.IGNORECASE) for pat in label_patterns):
            continue
        start = max(0, idx - 1)
        end = min(len(lines), idx + 1 + window)
        chunk = " ".join(lines[start:end])
        dates = _extract_dates_from_chunk(chunk)
        if dates:
            return dates[0]
    return ""


def _extract_dates(text: str, lines: List[str]) -> Tuple[str, str]:
    date_release = _find_date_near_label(lines, [r"дата\s+выпуск"])
    date_accept = _find_date_near_label(lines, [r"дата\s+при[её]м", r"свидетельство\s+о\s+при[её]мк"])

    if not date_release:
        for m in re.finditer(r"дата\s+выпуск[^\n]{0,90}", text, flags=re.IGNORECASE):
            dates = _extract_dates_from_chunk(m.group(0))
            if dates:
                date_release = dates[0]
                break

    if not date_accept:
        for m in re.finditer(r"дата\s+при[её]м[^\n]{0,90}", text, flags=re.IGNORECASE):
            dates = _extract_dates_from_chunk(m.group(0))
            if dates:
                date_accept = dates[0]
                break

    version_dates = _extract_version_dates(text)
    if date_release in version_dates:
        date_release = ""
    if date_accept in version_dates:
        date_accept = ""

    return date_release, date_accept


def _should_try_release_date_scan(raw_text: str) -> bool:
    t = (raw_text or "").lower().replace("ё", "е")
    if not t:
        return False
    if re.search(r"\bдата\s+(выпуск\w*|прием\w*|отк)\b", t):
        return True
    if "дата выпуска" in t or "дата приемки" in t:
        return True
    return False


def _extract_certificate_value(text: str, lines: List[str]) -> str:
    patterns = [
        r"\bСДС\.[A-ZА-Я0-9.\-\/]{6,}",
        r"\bЕАЭС\s*[A-ZА-Я0-9.\-\/]{6,}",
        r"\b(?:RU|РФ)\s*[A-ZА-Я0-9.\-\/]{5,}",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return re.sub(r"\s{2,}", " ", m.group(0)).strip(" .,:;")

    for idx, line in enumerate(lines):
        if not re.search(r"\bсертификат\w*\b", line, flags=re.IGNORECASE):
            continue
        chunk = " ".join(lines[idx: min(len(lines), idx + 4)])
        num = re.search(r"(?:№|N)\s*([A-ZА-Я0-9.\-\/]{5,})", chunk, flags=re.IGNORECASE)
        if num:
            return num.group(1).strip(" .,:;")

    by_label = _extract_value_after_label(lines, [r"\bсертификат\w*\b"])
    if not by_label:
        return ""
    s = by_label.strip(" .,:;")
    if len(s) > 80 or len(s.split()) > 5:
        return ""
    if not re.search(r"\d", s):
        return ""
    s_norm = s.lower().replace("ё", "е")
    if re.search(r"(соответств|лиценз|налич|услов|контролл|модул|шкаф|блок|перечн|документац)", s_norm):
        return ""
    if re.fullmatch(r"(соответстви[ея]|наличи[ея]|сертификат[а-я ]*)", s_norm, flags=re.IGNORECASE):
        return ""
    if not (
        re.search(r"[A-ZА-Я0-9]{2,}[.\-/][A-ZА-Я0-9.\-/]{2,}", s, flags=re.IGNORECASE)
        or re.search(r"(?:№|N)\s*[A-ZА-Я0-9.\-/]{5,}", s, flags=re.IGNORECASE)
    ):
        return ""
    return s


def _normalize_ocr_ru_token(value: str) -> str:
    s = (value or "").upper().replace("Ё", "Е")
    s = s.translate({
        ord("A"): "А",
        ord("B"): "В",
        ord("C"): "С",
        ord("E"): "Е",
        ord("H"): "Н",
        ord("K"): "К",
        ord("M"): "М",
        ord("O"): "О",
        ord("P"): "Р",
        ord("T"): "Т",
        ord("X"): "Х",
        ord("Y"): "У",
    })
    return re.sub(r"[^А-Я0-9]", "", s)


def _extract_blue_stamp_layer(img_rgb: Image.Image) -> Image.Image:
    r, g, b = img_rgb.split()
    rg_avg = ImageChops.blend(r, g, 0.5)
    blue = ImageChops.subtract(b, rg_avg)
    return ImageOps.autocontrast(blue)


def _date_from_roi(img: Image.Image, banned_dates: set, deadline: Optional[float] = None) -> str:
    prepared = [
        img,
        ImageEnhance.Contrast(img).enhance(1.8),
        img.point(lambda px: 255 if px > 155 else 0, mode="1").convert("L"),
    ]
    for variant in prepared:
        for psm in ("6", "7"):
            if deadline is not None and time.monotonic() > deadline:
                return ""
            try:
                txt = pytesseract.image_to_string(
                    variant,
                    lang=OCR_LANG,
                    config=f"--oem 1 --psm {psm}",
                    timeout=OCR_DATE_TESSERACT_TIMEOUT_SEC,
                )
            except Exception:
                continue
            for d in _extract_dates_from_chunk(txt):
                if d not in banned_dates:
                    return d
    return ""


def _extract_release_date_from_page_image(
    page_img: bytes,
    banned_dates: Optional[set] = None,
    deadline: Optional[float] = None,
) -> str:
    banned = set(banned_dates or set())
    try:
        img_rgb = Image.open(io.BytesIO(page_img)).convert("RGB")
    except Exception:
        return ""

    img = ImageOps.autocontrast(img_rgb.convert("L"))
    blue_layer = _extract_blue_stamp_layer(img_rgb)
    if img.width < 2200:
        scale = 2200 / float(max(1, img.width))
        img = img.resize((int(img.width * scale), int(img.height * scale)), Image.Resampling.LANCZOS)
        blue_layer = blue_layer.resize((int(blue_layer.width * scale), int(blue_layer.height * scale)), Image.Resampling.LANCZOS)
    img = ImageEnhance.Sharpness(img).enhance(1.7)
    blue_layer = ImageEnhance.Sharpness(blue_layer).enhance(1.5)

    label_rois: List[Tuple[int, int, int, int]] = []
    try:
        data = pytesseract.image_to_data(
            img,
            lang=OCR_LANG,
            config="--oem 1 --psm 6",
            output_type=pytesseract.Output.DICT,
            timeout=max(OCR_DATE_TESSERACT_TIMEOUT_SEC, 2),
        )
        words = data.get("text", [])
        left = data.get("left", [])
        top = data.get("top", [])
        heights = data.get("height", [])
        widths = data.get("width", [])
        normalized = [_normalize_ocr_ru_token(w) for w in words]

        for i, token in enumerate(normalized):
            if not token or not token.startswith("ДАТА"):
                continue
            nearby = normalized[i:i + 6]
            if not any(t.startswith("ВЫПУСК") for t in nearby):
                continue
            x = left[i]
            y = top[i]
            h = heights[i] if i < len(heights) else 20
            w = widths[i] if i < len(widths) else 60
            x1 = max(0, x - int(w * 0.8))
            y1 = max(0, y - int(h * 1.5))
            x2 = min(img.width, x + int(img.width * 0.55))
            y2 = min(img.height, y + int(img.height * 0.30))
            if x2 - x1 > 120 and y2 - y1 > 50:
                label_rois.append((x1, y1, x2, y2))
    except Exception:
        label_rois = []

    for box in label_rois:
        if deadline is not None and time.monotonic() > deadline:
            return ""
        for layer in (img, blue_layer):
            found = _date_from_roi(layer.crop(box), banned, deadline=deadline)
            if found:
                return found

    fallback_boxes = [
        (0, int(img.height * 0.30), int(img.width * 0.72), int(img.height * 0.78)),
        (0, int(img.height * 0.45), int(img.width * 0.72), int(img.height * 0.92)),
    ]
    for box in fallback_boxes:
        if deadline is not None and time.monotonic() > deadline:
            return ""
        for layer in (img, blue_layer):
            found = _date_from_roi(layer.crop(box), banned, deadline=deadline)
            if found:
                return found
    return ""


def _extract_release_date_from_pages(pages: List[bytes], banned_dates: Optional[set] = None) -> str:
    banned = set(banned_dates or set())
    deadline = time.monotonic() + OCR_RELEASE_SCAN_BUDGET_SEC
    for page_img in pages[:2]:
        if time.monotonic() > deadline:
            break
        found = _extract_release_date_from_page_image(page_img, banned_dates=banned, deadline=deadline)
        if found:
            return found
    return ""


def _extract_namenovanie(lines: List[str], full_text: str) -> str:
    for line in lines[:20]:
        if re.search(r"\bшкаф\b", line, flags=re.IGNORECASE):
            return line.strip()

    for idx, line in enumerate(lines[:150]):
        if not re.search(r"\bконтроллер\b", line, flags=re.IGNORECASE):
            continue
        if not re.search(r"\bмногофункциональ", line, flags=re.IGNORECASE):
            continue

        controller = line.strip(" .,:;")
        for j in range(idx + 1, min(len(lines), idx + 4)):
            if re.search(r"\b(мфк\s*[- ]?\d+)\b", lines[j], flags=re.IGNORECASE):
                controller = f"{controller} {lines[j].strip(' .,:;')}".strip()
                break

        module = ""
        for j in range(idx + 1, min(len(lines), idx + 10)):
            if re.search(r"\bмодул\w*\b", lines[j], flags=re.IGNORECASE):
                module = lines[j].strip(" .,:;")
                for k in range(j + 1, min(len(lines), j + 4)):
                    if re.search(r"\b(tcc\s*[a-z0-9-]*|tecon)\b", lines[k], flags=re.IGNORECASE):
                        module = f"{module} {lines[k].strip(' .,:;')}".strip()
                        break
                break

        if module:
            return f"{controller}; {module}".strip("; ")
        return controller

    for line in lines[:80]:
        if re.search(r"\bрозетк\w*\b", line, flags=re.IGNORECASE):
            if re.search(r"[A-Za-zА-Яа-я0-9]{4,}", line) and len(line) <= 140:
                return line.strip(" .,:;")

    for idx, line in enumerate(lines[:120]):
        if re.search(r"\bконтроллер\b", line, flags=re.IGNORECASE):
            if re.search(r"мфк\s*[- ]?\d+", line, flags=re.IGNORECASE):
                return line.strip(" .,:;")
            for j in range(idx + 1, min(len(lines), idx + 6)):
                if re.search(r"\b(мфк\s*[- ]?\d+|tcc\s*[a-z0-9-]*|tecon|текон)\b", lines[j], flags=re.IGNORECASE):
                    return f"{line}. {lines[j]}".strip(". ")

    for line in lines[:120]:
        if re.search(r"\b(мастер[- ]?модул|контроллер\s+многофункциональ)\b", line, flags=re.IGNORECASE):
            return line.strip(" .,:;")

    by_label = _extract_value_after_label(lines, [r"наименован\w*\s+издел", r"изделие"])
    if by_label and not re.search(r"\b(в соответствии|услови|предприят)\b", by_label, flags=re.IGNORECASE):
        if not re.fullmatch(r"[A-ZА-Я]{2,8}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?[A-ZА-Я]{0,3}", by_label):
            return by_label

    for line in lines[:120]:
        if len(line) > 140:
            continue
        if re.search(r"\b(розетк|модул|контроллер|шкаф|блок)\b", line, flags=re.IGNORECASE) and re.search(r"[A-Za-zА-Яа-я]+\d", line):
            return line.strip(" .,:;")

    for line in lines[:100]:
        if len(line) > 150:
            continue
        if re.search(r"\b(соответств|услови|предприят|контроль|эксплуатац)\b", line, flags=re.IGNORECASE):
            continue
        if re.search(r"\b(модул|контроллер|реле|панел|светильник|блок|шкаф)\b", line, flags=re.IGNORECASE):
            if re.search(r"\b(таблиц|гост|параметр|версия)\b", line, flags=re.IGNORECASE):
                continue
            return line.strip(" .,:;")

    m = re.search(r"(?:паспорт|руководство по эксплуатации)\s+([^\n]{4,140})", full_text, flags=re.IGNORECASE)
    if m:
        candidate = m.group(1).strip(" .,:;")
        if not re.fullmatch(r"[A-ZА-Я]{2,8}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?[A-ZА-Я]{0,3}", candidate):
            return candidate
    return ""


def _extract_structured_from_text(raw_text: str, source_name: str = "") -> Dict:
    lines = _clean_lines(raw_text)
    text = "\n".join(lines)
    text_lower = text.lower().replace("ё", "е")
    source_name_lower = (source_name or "").lower().replace("ё", "е")

    serials = _extract_serials(text)
    doc_code = _pick_document_code(text)
    non_passport_markers = [
        "служебная записка",
        "фактическое поступление",
        "рабочее место оператора",
        "критерии оценивания",
        "контекст задачи",
        "цель задачи",
        "хакатон",
        "экранная форма",
        "исходной таблицы",
        "автоматизация обработки паспортов",
        "условия участия",
        "оценивание",
        "кейс",
        "логистика",
        "разбор паспорта",
    ]
    source_non_passport_markers = [
        "задача",
        "логистика",
        "разбор",
        "экранная форма",
        "кейс",
        "служебная",
    ]
    has_passport_word = bool(re.search(r"\bпаспорт\b", text_lower)) or bool(re.search(r"\bпаспорт\b", source_name_lower))
    has_manual_word = ("руководство по эксплуатации" in text_lower) or ("руковод" in source_name_lower)
    signal_count = 0
    if doc_code:
        signal_count += 2
    if serials:
        signal_count += 2
    if re.search(r"\b(заводск\w*\s+номер|код\s+заказ|гарант\w*|срок\s+служб\w*|сертификат)\b", text_lower, flags=re.IGNORECASE):
        signal_count += 1
    if re.search(r"\b(контроллер|модул|блок|реле|шкаф|розетк|панел)\b", text_lower, flags=re.IGNORECASE):
        signal_count += 1
    has_equipment_signals = signal_count >= 2

    is_non_passport = any(marker in text_lower for marker in non_passport_markers) or any(
        marker in source_name_lower for marker in source_non_passport_markers
    )
    allow_draft_passport = (has_passport_word or has_manual_word) and has_equipment_signals

    doc_type = "unknown"
    if "перечень документации" in text_lower or ("шкаф" in text_lower and "зав." in text_lower):
        doc_type = "cabinet_list"
    elif is_non_passport and not allow_draft_passport:
        doc_type = "unknown"
    elif has_passport_word or has_manual_word:
        if "tcc8l" in text_lower or "мфк1500" in text_lower or len(serials) > 1:
            doc_type = "group_passport"
        else:
            doc_type = "single_passport"

    name = _extract_namenovanie(lines, text)
    kod_zakaza = _extract_value_after_label(lines, [r"код\s+заказ\w*"])
    if kod_zakaza:
        kz_norm = _normalize_serial_token(kod_zakaza)
        serials = [s for s in serials if _normalize_serial_token(s) != kz_norm]
        if doc_type == "group_passport" and len(serials) <= 1 and "tcc8l" not in text_lower and "мфк1500" not in text_lower:
            doc_type = "single_passport"
    garantia = _extract_value_after_label(lines, [r"\bгарант\w*\b"])
    srok = _extract_value_after_label(lines, [r"\bсрок\s+служб\w*\b"])
    sert = _extract_certificate_value(text, lines)
    date_release, date_accept = _extract_dates(text, lines)

    manufacturer = ""
    manufacturer_match = re.search(r"\b(?:АО|ПАО|ООО|ЗАО)\s*[\"«][^\"»\n]{2,80}[\"»]", text, flags=re.IGNORECASE)
    if manufacturer_match:
        manufacturer = manufacturer_match.group(0).strip()
    else:
        manufacturer_match = re.search(r"\b(?:АО|ПАО|ООО|ЗАО)\s+[А-ЯA-Z][^\n,]{2,80}", text, flags=re.IGNORECASE)
        if manufacturer_match:
            manufacturer = manufacturer_match.group(0).strip()

    address = ""
    for line in lines:
        if len(line) < 10 or len(line) > 180:
            continue
        if "," not in line:
            continue
        if re.search(r"(россия|рф|г\.\s*[а-я]|ул\.|улиц|просп|пр-т|д\.|дом|корп|стр\.)", line, flags=re.IGNORECASE) and re.search(r"\d", line):
            if len(re.findall(r"\w+", line)) > 24:
                continue
            address = line
            break

    contacts = []
    contacts.extend(re.findall(r"(?:\+7|8)\s*\(?\d{3,4}\)?[\s\-]?\d{2,3}[\s\-]?\d{2}[\s\-]?\d{2}", text))
    contacts.extend(re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.IGNORECASE))

    extracted = {
        "document_type": doc_type,
        "naimenovanie": name,
        "kod_dokumenta": doc_code,
        "proizvoditel": manufacturer,
        "adres": address,
        "kontakty": ", ".join(dict.fromkeys(contacts)) if contacts else "",
        "garantia": garantia,
        "srok_sluzhby": srok,
        "sertifikat": sert,
        "kod_zakaza": kod_zakaza,
        "normativnye_dok": _extract_normative_docs(text),
        "zavodskie_nomera": serials,
        "data_vypuska": date_release,
        "data_priemki": date_accept,
        "_non_passport": is_non_passport,
        "_allow_draft_passport": allow_draft_passport,
    }
    if extracted["garantia"] and not _is_duration_like(extracted["garantia"]):
        extracted["garantia"] = ""
    if extracted["srok_sluzhby"] and not _is_duration_like(extracted["srok_sluzhby"]):
        extracted["srok_sluzhby"] = ""
    if extracted["sertifikat"] and (len(extracted["sertifikat"]) > 80 or not re.search(r"[A-ZА-Я0-9]", extracted["sertifikat"], flags=re.IGNORECASE)):
        extracted["sertifikat"] = ""
    if extracted["kod_zakaza"] and len(extracted["kod_zakaza"]) > 40:
        extracted["kod_zakaza"] = ""

    return extracted


def parse_cabinet_document(raw_text: str) -> Dict:
    lines = _clean_lines(raw_text)
    text = "\n".join(lines)
    result = {
        "shkaf_naim": "",
        "shkaf_kod": "",
        "shkaf_zav_nomer": "",
        "pozicii": [],
    }

    for line in lines[:20]:
        if re.search(r"\bшкаф\b", line, flags=re.IGNORECASE):
            result["shkaf_naim"] = line
            code_match = re.search(r"\(([^)]+)\)", line)
            if code_match:
                result["shkaf_kod"] = code_match.group(1).strip()
            break

    serial_header_idx = next((i for i, l in enumerate(lines) if re.search(r"заводск\w*\s+номер", l, flags=re.IGNORECASE)), -1)
    serial_source = "\n".join(lines[serial_header_idx:]) if serial_header_idx >= 0 else text

    zav_match = re.search(r"зав\.\s*№\s*([A-ZА-Я0-9\-]+)", serial_source, flags=re.IGNORECASE)
    if zav_match:
        result["shkaf_zav_nomer"] = zav_match.group(1).strip()

    doc_code_pattern = re.compile(r"\b[А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,4}(?:-\d{2,3})?(?:\s*[А-ЯA-Z]{1,3})?\b")
    serial_pattern = re.compile(r"\b(?:\d{9,14}|[A-ZА-Я0-9]{5,20}|б/н|Б/Н)\b")

    items = []
    for i, line in enumerate(lines):
        for m in serial_pattern.finditer(line):
            serial = m.group(0).strip()
            if serial.lower() != "б/н" and not _is_probable_serial(serial):
                continue

            name_parts = []
            for j in range(i - 1, max(-1, i - 4), -1):
                candidate = lines[j]
                if re.search(r"(перечень|заводской|страниц|листов|сертификат|паспорт\b)", candidate, flags=re.IGNORECASE):
                    continue
                if re.fullmatch(r"\d+[.)]?", candidate):
                    continue
                if len(candidate) < 4:
                    continue
                name_parts.insert(0, candidate)
                if len(" ".join(name_parts)) >= 90:
                    break

            context_lines = lines[max(0, i - 4):min(len(lines), i + 3)]
            context_text = " ".join(context_lines)
            doc_match = doc_code_pattern.search(context_text)
            if not doc_match:
                passport_line = next(
                    (ln for ln in reversed(context_lines) if re.search(r"\bпаспорт\b", ln, flags=re.IGNORECASE)),
                    "",
                )
                doc_match = doc_code_pattern.search(passport_line) if passport_line else None
            doc_code = re.sub(r"\s+", "", doc_match.group(0)).strip() if doc_match else ""

            items.append({
                "naimenovanie": re.sub(r"\s{2,}", " ", " ".join(name_parts)).strip() or f"Позиция {len(items) + 1}",
                "zavodskoy_nomer": serial,
                "oboznachenie_dok": doc_code,
            })

    seen = set()
    for item in items:
        key = (item["zavodskoy_nomer"], item["oboznachenie_dok"], item["naimenovanie"])
        if key in seen:
            continue
        seen.add(key)
        result["pozicii"].append({
            "nomer": len(result["pozicii"]) + 1,
            **item,
        })

    return result


def _has_nonempty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, list):
        return any(str(v).strip() for v in value)
    return bool(str(value).strip())


def _sanitize_extracted_payload(payload: Dict) -> Dict:
    data = dict(payload or {})

    def _clean_text(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    allowed_doc_types = {"single_passport", "group_passport", "cabinet_list", "unknown"}
    doc_type = _clean_text(data.get("document_type")).lower()
    data["document_type"] = doc_type if doc_type in allowed_doc_types else "unknown"

    serials_raw = data.get("zavodskie_nomera") or []
    if not isinstance(serials_raw, list):
        serials_raw = [serials_raw]
    allow_short_numeric_serial = doc_type in {"group_passport", "cabinet_list"}
    data["zavodskie_nomera"] = [
        s for s in normalize_serials(serials_raw)
        if _is_probable_serial(s) or (allow_short_numeric_serial and re.fullmatch(r"\d{7}", s))
    ]

    norm_raw = data.get("normativnye_dok") or []
    if not isinstance(norm_raw, list):
        norm_raw = [norm_raw]
    norm_clean: List[str] = []
    for item in norm_raw:
        s = _clean_text(item)
        if not s or _is_placeholder_scalar(s):
            continue
        if re.search(r"\b(гост|ту|тр\s*тс|iso|iec)\b", s, flags=re.IGNORECASE) or _pick_document_code(s):
            norm_clean.append(s)
    data["normativnye_dok"] = list(dict.fromkeys(norm_clean))

    comp_raw = data.get("komplektnost") or []
    if not isinstance(comp_raw, list):
        comp_raw = [comp_raw]
    data["komplektnost"] = [
        _clean_text(x) for x in comp_raw
        if _clean_text(x) and not _is_placeholder_scalar(x)
    ]

    for key in (
        "naimenovanie",
        "proizvoditel",
        "adres",
        "kontakty",
        "kod_zakaza",
        "garantia",
        "srok_sluzhby",
        "sertifikat",
        "data_vypuska",
        "data_priemki",
        "kod_dokumenta",
    ):
        if key in data:
            data[key] = _clean_text(data.get(key))
            if _is_placeholder_scalar(data[key]):
                data[key] = ""

    if data.get("kod_dokumenta"):
        normalized_code = _pick_document_code(data["kod_dokumenta"])
        data["kod_dokumenta"] = normalized_code or ""

    if data.get("kod_zakaza"):
        kz = data["kod_zakaza"].strip(" .,:;")
        if ":" in kz or ";" in kz:
            candidates = [
                t for t in re.findall(r"[A-Za-zА-Яа-я0-9][A-Za-zА-Яа-я0-9\-./]{2,}", kz)
                if re.search(r"\d", t)
            ]
            if candidates:
                kz = max(candidates, key=len)
        kz = kz.strip(" .,:;")
        if re.fullmatch(r"[A-Za-zА-Яа-я0-9./\-]{3,40}", kz) and re.search(r"\d", kz):
            kz = kz.upper().translate(CYR_TO_LAT_LOOKALIKE)
        if len(kz) < 3 or re.fullmatch(r"[0-9]{1,2}", kz):
            kz = ""
        data["kod_zakaza"] = kz

    if data.get("data_vypuska"):
        dates = _extract_dates_from_chunk(data["data_vypuska"])
        data["data_vypuska"] = dates[0] if dates else ""
    if data.get("data_priemki"):
        dates = _extract_dates_from_chunk(data["data_priemki"])
        data["data_priemki"] = dates[0] if dates else ""

    if data.get("garantia") and not _is_duration_like(data["garantia"]):
        data["garantia"] = ""
    if data.get("srok_sluzhby") and not _is_duration_like(data["srok_sluzhby"]):
        data["srok_sluzhby"] = ""

    if data.get("naimenovanie"):
        nm = data["naimenovanie"]
        has_letters = len(re.findall(r"[A-Za-zА-Яа-я]", nm)) >= 2
        has_equipment_signal = bool(
            re.search(
                r"\b(розетк|модул|контрол|шкаф|блок|панел|реле|мфк|tecon|tcc|оптидин|optidin|master)\b",
                nm,
                flags=re.IGNORECASE,
            )
            or re.search(r"[A-Za-zА-Яа-я]+\d{2,}", nm)
            or re.search(r"\d{2,}[A-Za-zА-Яа-я]+", nm)
        )
        if len(nm) < 4 or not has_letters or not has_equipment_signal:
            data["naimenovanie"] = ""
    if data.get("proizvoditel"):
        pr = data["proizvoditel"]
        pr_upper = pr.upper()
        has_org_marker = bool(re.search(r"\b(?:АО|ПАО|ООО|ЗАО)\b", pr_upper))
        has_vendor_marker = bool(re.search(r"\b(КЭАЗ|TREI|ТРЭИ|TECON|OPTIDIN|OPTI)\b", pr_upper))
        if len(pr) < 3 or len(re.findall(r"[A-Za-zА-Яа-я]", pr)) < 2:
            data["proizvoditel"] = ""
        elif not has_org_marker and not has_vendor_marker and (len(pr) < 6 or not re.search(r"[А-Яа-я]{4,}", pr)):
            data["proizvoditel"] = ""

    if data.get("adres"):
        ad = data["adres"]
        anchor = re.search(r"\b(россия|рф|г\.)\b", ad, flags=re.IGNORECASE)
        if anchor and anchor.start() > 0:
            prefix = ad[:anchor.start()]
            if len(prefix.strip()) <= 24:
                ad = ad[anchor.start():].lstrip(" ,;:-")
        if len(ad) < 8 or len(ad) > 180 or ("," not in ad and not re.search(r"\b(ул|улиц|г\.|город|д\.|дом)\b", ad, flags=re.IGNORECASE)):
            data["adres"] = ""
        else:
            data["adres"] = ad

    if data.get("kontakty"):
        c = data["kontakty"]
        phones = re.findall(r"(?:\+7|8)\s*\(?\d{3,4}\)?[\s\-]?\d{2,3}[\s\-]?\d{2}[\s\-]?\d{2}", c)
        emails = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", c, flags=re.IGNORECASE)
        uniq = list(dict.fromkeys([*phones, *emails]))
        data["kontakty"] = ", ".join(uniq) if uniq else ""

    if data.get("sertifikat"):
        cert = data["sertifikat"]
        if len(cert) < 5 or len(cert) > 80 or not re.search(r"\d", cert):
            data["sertifikat"] = ""

    for key in (
        "naimenovanie",
        "kod_dokumenta",
        "proizvoditel",
        "adres",
        "kontakty",
        "garantia",
        "srok_sluzhby",
        "sertifikat",
        "kod_zakaza",
        "data_vypuska",
        "data_priemki",
    ):
        if not data.get(key):
            data[key] = None

    return data


def _quality_assessment(data: Dict) -> Tuple[int, List[str], bool]:
    doc_type = str(data.get("document_type") or "unknown").strip().lower()
    score = 0
    missing_fields: List[str] = []

    weighted_fields = [
        ("naimenovanie", "Наименование", 24),
        ("kod_dokumenta", "Код документа", 22),
        ("proizvoditel", "Производитель", 12),
        ("zavodskie_nomera", "Серийные номера", 18),
        ("data_vypuska", "Дата выпуска", 10),
        ("kod_zakaza", "Код заказа", 6),
        ("sertifikat", "Сертификат", 4),
        ("adres", "Адрес", 2),
        ("kontakty", "Контакты", 2),
    ]
    for key, label, weight in weighted_fields:
        if _has_nonempty(data.get(key)):
            score += weight
        elif key in {"naimenovanie", "kod_dokumenta"}:
            missing_fields.append(label)

    if not _has_nonempty(data.get("zavodskie_nomera")) and not _has_nonempty(data.get("data_vypuska")):
        missing_fields.append("Серийный номер или дата выпуска")

    if doc_type == "unknown":
        score = max(0, score - 10)
    elif doc_type == "cabinet_list":
        score = max(score, 45 if _has_nonempty(data.get("naimenovanie")) else score)
        missing_fields = [m for m in missing_fields if m != "Серийный номер или дата выпуска"]

    dedup_missing: List[str] = []
    seen = set()
    for item in missing_fields:
        if item not in seen:
            seen.add(item)
            dedup_missing.append(item)

    score = int(max(0, min(100, round(score))))
    needs_review = score < 60 or bool(dedup_missing)
    return score, dedup_missing, needs_review


def final_cleanup(data: Dict, raw_text: str = "", source_name: str = "") -> Dict:
    data = _sanitize_extracted_payload(data or {})
    heur = _extract_structured_from_text(raw_text, source_name=source_name)

    merge_keys = [
        "document_type",
        "naimenovanie",
        "kod_dokumenta",
        "proizvoditel",
        "adres",
        "kontakty",
        "garantia",
        "srok_sluzhby",
        "sertifikat",
        "kod_zakaza",
        "normativnye_dok",
        "zavodskie_nomera",
        "data_vypuska",
        "data_priemki",
    ]
    for key in merge_keys:
        value = heur.get(key)
        if value:
            data[key] = value

    if data.get("naimenovanie"):
        name = re.sub(
            r"\b(паспорт|руководство|эксплуатации|свидетельство|приемке|упаковывании)\b",
            "",
            str(data["naimenovanie"]),
            flags=re.IGNORECASE,
        )
        name = re.sub(r"^[A-Za-zА-Яа-я]{1,3}\s*:\s*", "", name)
        data["naimenovanie"] = re.sub(r"\s{2,}", " ", name).strip(" ,.-[]")

    data["zavodskie_nomera"] = normalize_serials(data.get("zavodskie_nomera"))
    data["normativnye_dok"] = [x for x in (data.get("normativnye_dok") or []) if str(x).strip()]
    data["komplektnost"] = [x for x in (data.get("komplektnost") or []) if str(x).strip()]

    fb = regex_fallbacks(raw_text)
    for k, v in fb.items():
        if k == "zavodskie_nomera":
            merged = normalize_serials((data.get(k) or []) + v)
            data[k] = merged
        elif not data.get(k):
            data[k] = v

    data = _sanitize_extracted_payload(data)

    data["zavodskie_nomera"] = [s for s in data.get("zavodskie_nomera", []) if not re.fullmatch(r"\d{1,5}", s)]
    if data.get("kod_zakaza"):
        kz_norm = _normalize_serial_token(str(data.get("kod_zakaza", "")))
        data["zavodskie_nomera"] = [
            s for s in data.get("zavodskie_nomera", [])
            if _normalize_serial_token(s) != kz_norm
        ]
    doc_type = (data.get("document_type") or "").strip().lower()
    text_lower = (raw_text or "").lower().replace("ё", "е")
    labeled_serials = _extract_labeled_factory_serials(raw_text)
    has_group_markers = bool(re.search(r"\b(tcc8l|мфк1500|перечень\s+документации)\b", text_lower))
    if doc_type in {"single_passport", "group_passport"} and len(labeled_serials) == 1 and not has_group_markers:
        data["zavodskie_nomera"] = [labeled_serials[0]]
        if doc_type == "group_passport":
            data["document_type"] = "single_passport"
            doc_type = "single_passport"

    if doc_type == "unknown":
        if heur.get("_allow_draft_passport"):
            data["document_type"] = "single_passport"
            doc_type = "single_passport"
        elif not data.get("naimenovanie") and not data.get("kod_dokumenta"):
            data["zavodskie_nomera"] = []
            for key in ("proizvoditel", "adres", "kontakty", "garantia", "srok_sluzhby", "sertifikat", "kod_zakaza", "data_vypuska", "data_priemki"):
                data[key] = None
            data["normativnye_dok"] = []
    elif doc_type == "cabinet_list":
        for key in ("garantia", "srok_sluzhby", "sertifikat", "data_vypuska", "data_priemki"):
            data[key] = None
        cabinet = parse_cabinet_document(raw_text or "")
        serials_from_positions = [
            str(p.get("zavodskoy_nomer", "")).strip()
            for p in (cabinet.get("pozicii") or [])
            if str(p.get("zavodskoy_nomer", "")).strip() and str(p.get("zavodskoy_nomer", "")).strip().lower() != "б/н"
        ]
        if serials_from_positions:
            data["zavodskie_nomera"] = normalize_serials((data.get("zavodskie_nomera") or []) + serials_from_positions)

    if doc_type == "single_passport":
        serials = normalize_serials(data.get("zavodskie_nomera"))
        if len(serials) > 1:
            by_label = _extract_labeled_factory_serials(raw_text)
            if by_label:
                label_norm = {_normalize_serial_token(x) for x in by_label}
                picked = [s for s in serials if _normalize_serial_token(s) in label_norm]
                data["zavodskie_nomera"] = [picked[0] if picked else by_label[0]]
            else:
                data["zavodskie_nomera"] = [max(serials, key=_serial_plausibility_score)]

    if data.get("naimenovanie"):
        name = str(data.get("naimenovanie", "")).strip()
        if re.fullmatch(r"контроллер\s+многофункциональн(ый|ого)?", name, flags=re.IGNORECASE):
            data["naimenovanie"] = None

    if data.get("adres"):
        addr = str(data["adres"]).strip()
        if len(addr) > 140 or len(re.findall(r"\w+", addr)) > 24:
            data["adres"] = None

    if doc_type in {"group_passport", "cabinet_list"}:
        data["tip_pasporta"] = "group"
    elif data.get("zavodskie_nomera"):
        data["tip_pasporta"] = "individual"
    else:
        data["tip_pasporta"] = "no_serial"

    quality_score, missing_fields, needs_review = _quality_assessment(data)
    data["quality_score"] = quality_score
    data["missing_fields"] = missing_fields
    data["needs_review"] = needs_review

    return data


def _build_record_checklist(data: Dict) -> List[Dict[str, Any]]:
    doc_type = str(data.get("document_type") or "unknown").strip().lower()
    serials = normalize_serials(data.get("zavodskie_nomera"))
    checklist: List[Dict[str, Any]] = []

    def add(code: str, title: str, ok: bool, details: str = ""):
        checklist.append({
            "code": code,
            "title": title,
            "ok": bool(ok),
            "details": details or "",
        })

    add("doc_type", "Тип документа определен", doc_type in {"single_passport", "group_passport", "cabinet_list"}, doc_type or "unknown")
    add("name", "Есть наименование", bool(data.get("naimenovanie")), str(data.get("naimenovanie") or ""))
    add("doc_code", "Есть код документа", bool(data.get("kod_dokumenta")), str(data.get("kod_dokumenta") or ""))
    add("serial_or_date", "Есть серийный номер или дата выпуска", bool(serials) or bool(data.get("data_vypuska")), f"serials={len(serials)}, date={data.get('data_vypuska') or ''}")
    add("not_review", "Не требует ручной проверки", not bool(data.get("needs_review")), f"quality={data.get('quality_score')}")

    if doc_type in {"single_passport", "group_passport"}:
        add("manufacturer", "Есть производитель", bool(data.get("proizvoditel")), str(data.get("proizvoditel") or ""))
        add("order_or_serial", "Есть код заказа или серийный номер", bool(data.get("kod_zakaza")) or bool(serials), f"order={data.get('kod_zakaza') or ''}; serials={len(serials)}")
    elif doc_type == "cabinet_list":
        add("cabinet_mode", "Документ пригоден для режима шкафа", True, "cabinet_list")

    return checklist


def _snippet_from_text(text: str, start: int, end: int, radius: int = 80) -> str:
    if not text:
        return ""
    s = max(0, start - radius)
    e = min(len(text), end + radius)
    chunk = re.sub(r"\s+", " ", text[s:e]).strip()
    if len(chunk) > 220:
        chunk = chunk[:220].rstrip() + "..."
    return chunk


def _find_value_evidence(value: str, per_page_texts: List[str], label_pattern: str = "") -> Dict[str, Any]:
    needle = str(value or "").strip()
    if not needle:
        return {}

    pattern = re.escape(needle)
    pattern = pattern.replace(r"\ ", r"\s*")
    for idx, page_text in enumerate(per_page_texts):
        if not page_text:
            continue
        m = re.search(pattern, page_text, flags=re.IGNORECASE)
        if m:
            return {
                "page": idx + 1,
                "snippet": _snippet_from_text(page_text, m.start(), m.end()),
            }

    if label_pattern:
        for idx, page_text in enumerate(per_page_texts):
            if not page_text:
                continue
            lm = re.search(label_pattern, page_text, flags=re.IGNORECASE)
            if lm:
                return {
                    "page": idx + 1,
                    "snippet": _snippet_from_text(page_text, lm.start(), lm.end()),
                }

    return {}


def _build_field_evidence(data: Dict, per_page_texts: List[str]) -> Dict[str, Any]:
    if not per_page_texts:
        return {}

    label_map = {
        "naimenovanie": r"наименован\w*|издел",
        "kod_dokumenta": r"код\s+документ|паспорт|обозначение",
        "kod_zakaza": r"код\s+заказ\w*",
        "data_vypuska": r"дата\s+выпуск\w*",
        "data_priemki": r"дата\s+при[её]м\w*|отк",
        "garantia": r"гарант\w*",
        "srok_sluzhby": r"срок\s+служб\w*",
        "sertifikat": r"сертификат\w*",
        "proizvoditel": r"(?:АО|ПАО|ООО|ЗАО)|производител\w*",
        "adres": r"адрес|ул\.|город|г\.",
        "kontakty": r"контакт|тел\.|e-?mail",
    }

    evidence: Dict[str, Any] = {}
    for field in (
        "naimenovanie",
        "kod_dokumenta",
        "kod_zakaza",
        "data_vypuska",
        "data_priemki",
        "garantia",
        "srok_sluzhby",
        "sertifikat",
        "proizvoditel",
        "adres",
        "kontakty",
    ):
        val = data.get(field)
        if not val:
            continue
        ev = _find_value_evidence(str(val), per_page_texts, label_pattern=label_map.get(field, ""))
        if ev:
            evidence[field] = {"value": str(val), **ev}

    serial_evidence = []
    for sn in normalize_serials(data.get("zavodskie_nomera")):
        ev = _find_value_evidence(sn, per_page_texts, label_pattern=r"завод\w*\s*номер|серийн\w*\s*номер")
        if ev:
            serial_evidence.append({"value": sn, **ev})
        else:
            serial_evidence.append({"value": sn})
    if serial_evidence:
        evidence["zavodskie_nomera"] = serial_evidence

    return evidence


def _build_field_risks(data: Dict, evidence: Optional[Dict[str, Any]] = None) -> Dict[str, Dict[str, str]]:
    doc_type = str(data.get("document_type") or "unknown").strip().lower()
    serials = normalize_serials(data.get("zavodskie_nomera"))
    evidence = evidence if isinstance(evidence, dict) else {}
    risks: Dict[str, Dict[str, str]] = {}
    rank = {"low": 1, "medium": 2, "high": 3}

    def add(field: str, level: str, reason: str):
        level = (level or "low").strip().lower()
        if level not in rank:
            level = "low"
        current = risks.get(field)
        if current and rank.get(current.get("level", "low"), 1) >= rank[level]:
            return
        risks[field] = {"level": level, "reason": reason[:220]}

    critical = ["naimenovanie", "kod_dokumenta"]
    if doc_type in {"single_passport", "group_passport"}:
        critical.append("proizvoditel")
    for field in critical:
        if not str(data.get(field) or "").strip():
            add(field, "high", "Ключевое поле не извлечено")

    if doc_type == "single_passport" and not serials:
        add("zavodskie_nomera", "high", "Для индивидуального паспорта отсутствует серийный номер")
    elif doc_type == "group_passport" and len(serials) < 2:
        add("zavodskie_nomera", "medium", "Для группового паспорта найдено мало серийных номеров")
    elif doc_type in {"single_passport", "group_passport"} and not data.get("data_vypuska"):
        add("data_vypuska", "low", "Дата выпуска не найдена")

    for date_field in ("data_vypuska", "data_priemki"):
        val = str(data.get(date_field) or "").strip()
        if val and not _is_date_like(val):
            add(date_field, "medium", "Нетипичный формат даты")

    if str(data.get("garantia") or "").strip() and not _is_duration_like(str(data.get("garantia") or "")):
        add("garantia", "medium", "Гарантия не похожа на срок/период")
    if str(data.get("srok_sluzhby") or "").strip() and not _is_duration_like(str(data.get("srok_sluzhby") or "")):
        add("srok_sluzhby", "medium", "Срок службы не похож на срок/период")

    for field in ("naimenovanie", "kod_dokumenta", "kod_zakaza", "data_vypuska", "data_priemki", "proizvoditel", "adres", "kontakty"):
        if str(data.get(field) or "").strip() and field not in evidence:
            add(field, "low", "Для значения не найден прямой источник в тексте")

    if serials and "zavodskie_nomera" not in evidence:
        add("zavodskie_nomera", "low", "Серийные номера извлечены, но источник не зафиксирован")

    return risks


def _as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    s = str(value).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _normalize_eval_scalar(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("ё", "е").lower()).strip()


def _compare_eval_value(expected: Any, predicted: Any) -> bool:
    if isinstance(expected, list):
        exp = sorted(_normalize_eval_scalar(x) for x in expected if _normalize_eval_scalar(x))
        pred = sorted(_normalize_eval_scalar(x) for x in (predicted or []) if _normalize_eval_scalar(x))
        return exp == pred
    return _normalize_eval_scalar(expected) == _normalize_eval_scalar(predicted)


def _extract_document_payload(
    fbytes: bytes,
    filename: str,
    allow_llm: Optional[bool] = None,
    scan_release_date: bool = True,
) -> Dict[str, Any]:
    if filename.lower().endswith(".pdf"):
        pages, embedded_text_blocks = pdf_to_images_and_text(fbytes)
    else:
        pages, embedded_text_blocks = [fbytes], [""]

    provider, active_model = resolve_llm_backend()
    run_ocr = True
    ocr_texts = collect_ocr_texts(pages, embedded_text_blocks, enabled=run_ocr)
    per_page_texts: List[str] = []
    for idx in range(len(pages)):
        embedded = (embedded_text_blocks[idx] if idx < len(embedded_text_blocks) else "").strip()
        ocr = (ocr_texts[idx] if idx < len(ocr_texts) else "").strip()
        per_page_texts.append("\n".join(part for part in (embedded, ocr) if part).strip())

    full_text = "\n".join(part for part in per_page_texts if part).strip()
    heur_preview = final_cleanup({}, raw_text=full_text, source_name=filename)

    llm_images: List[str] = []
    llm_data = {}
    llm_error = ""
    skip_by_content = _should_skip_llm_for_file(filename, full_text)
    skip_by_quality = _has_strong_heuristic_result(heur_preview)
    llm_enabled = ENABLE_LLM if allow_llm is None else bool(allow_llm)
    llm_available = llm_enabled and provider != "disabled"

    if llm_available and not skip_by_content and not skip_by_quality:
        llm_images = prepare_llm_images_b64(pages)
        llm_input_text = full_text
        if int(heur_preview.get("quality_score") or 0) < 25:
            llm_input_text = _prioritize_ocr_for_llm(full_text, max(1200, LLM_FALLBACK_TEXT_LIMIT))
        try:
            llm_data = run_llm_extraction(llm_images, llm_input_text, provider=provider, model_name=active_model)
        except requests.RequestException as e:
            llm_error = str(e)
            log.info("LLM extraction failed, fallback to OCR+rules only: %s", llm_error)
    elif llm_available:
        if skip_by_content:
            llm_error = "LLM skipped: likely non-passport/service input"
        elif skip_by_quality:
            llm_error = "LLM skipped: OCR+rules confidence is already high"
        else:
            llm_error = "LLM skipped"
    else:
        if llm_enabled:
            llm_error = "LLM unavailable: Ollama is not reachable"
        else:
            llm_error = "LLM disabled (ENABLE_LLM=0)"

    final_result = heur_preview if not llm_data else final_cleanup(llm_data, raw_text=full_text, source_name=filename)
    if (
        scan_release_date
        and
        final_result.get("document_type") != "unknown"
        and not final_result.get("data_vypuska")
        and _should_try_release_date_scan(full_text)
    ):
        version_dates = _extract_version_dates(full_text)
        release_date = _extract_release_date_from_pages(pages, banned_dates=version_dates)
        if release_date:
            final_result["data_vypuska"] = release_date

    evidence_map = _build_field_evidence(final_result, per_page_texts)
    final_result["_evidence"] = evidence_map
    final_result["_field_risks"] = _build_field_risks(final_result, evidence=evidence_map)
    final_result["_checklist"] = _build_record_checklist(final_result)
    final_result["_meta"] = {
        "provider": provider,
        "model": active_model,
        "pages_processed": len(pages),
        "images_sent_to_llm": len(llm_images),
        "ocr_enabled": run_ocr,
        "llm_error": llm_error,
        "llm_used": bool(llm_data),
        "scan_release_date": bool(scan_release_date),
    }
    return final_result



@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/extract", methods=["POST"])
def extract():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file"}), 400

    try:
        result = _extract_document_payload(file.read(), file.filename or "uploaded_file")
        return jsonify(result)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/parse_cabinet", methods=["POST"])
def parse_cabinet():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file"}), 400

    try:
        fbytes = file.read()
        if file.filename.lower().endswith(".pdf"):
            pages, embedded_text_blocks = pdf_to_images_and_text(fbytes)
        else:
            pages, embedded_text_blocks = [fbytes], [""]

        ocr_texts = collect_ocr_texts(pages, embedded_text_blocks, enabled=True)
        full_text = "\n".join(part for part in ("\n".join(embedded_text_blocks), "\n".join(ocr_texts)) if part).strip()
        result = parse_cabinet_document(full_text)
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/barcode", methods=["POST"])
def api_barcode():
    val = (request.json or {}).get("value", "").strip()
    if not val:
        return jsonify({"error": "Empty"}), 400
    bc_class = barcode.get_barcode_class("code128")
    buf = io.BytesIO()
    clean_val = "".join(re.findall(r"[A-Z0-9\-]", val.upper()))
    bc_class(clean_val[:20] or "0000", writer=ImageWriter()).write(buf)
    return jsonify({"barcode": base64.b64encode(buf.getvalue()).decode(), "encoded_value": clean_val})


@app.route("/api/preview", methods=["POST"])
def preview():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file"}), 400
    fbytes = file.read()
    images = []
    if file.filename.lower().endswith(".pdf"):
        doc = fitz.open(stream=fbytes, filetype="pdf")
        for page in doc:
            img = page.get_pixmap(matrix=fitz.Matrix(PREVIEW_RENDER_SCALE, PREVIEW_RENDER_SCALE)).tobytes("png")
            images.append(base64.b64encode(img).decode())
    else:
        images.append(base64.b64encode(fbytes).decode())
    return jsonify({
        "images": images,
        "image": images[0] if images else "",
        "page_count": len(images),
    })


@app.route("/api/meta", methods=["GET"])
def meta():
    provider, active_model = resolve_llm_backend()
    return jsonify({
        "active_provider": provider,
        "active_model": active_model,
        "model_candidates": MODEL_CANDIDATES if provider == "ollama" else [],
        "max_llm_images": MAX_LLM_IMAGES,
        "ocr_workers": OCR_WORKERS,
        "enable_llm": ENABLE_LLM,
        "llm_timeout_sec": LLM_REQUEST_TIMEOUT,
        "llm_connect_timeout_sec": LLM_CONNECT_TIMEOUT,
        "llm_max_total_sec": LLM_MAX_TOTAL_SEC,
        "llm_model_failover_tries": LLM_MODEL_FAILOVER_TRIES,
        "llm_primary_max_images": LLM_PRIMARY_MAX_IMAGES,
        "ollama_reachable": bool(_ollama_health_cache.get("alive")),
        "ollama_last_error": _ollama_health_cache.get("last_error", ""),
        "ocr_psms": OCR_PSMS,
        "ocr_tesseract_timeout_sec": OCR_TESSERACT_TIMEOUT_SEC,
        "ocr_date_tesseract_timeout_sec": OCR_DATE_TESSERACT_TIMEOUT_SEC,
        "ocr_release_scan_budget_sec": OCR_RELEASE_SCAN_BUDGET_SEC,
        "registry_b64_max": REGISTRY_B64_MAX,
        "registry_max_images": REGISTRY_MAX_IMAGES,
        "feedback_file": str(_feedback_path()),
    })


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _registry_path() -> Path:
    configured = Path(REGISTRY_STATE_FILE)
    return configured if configured.is_absolute() else (_project_root() / configured)


def _feedback_path() -> Path:
    configured = Path(FEEDBACK_FILE)
    return configured if configured.is_absolute() else (_project_root() / configured)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _clean_registry_scalar(value: Any, max_len: int = 220) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:max_len]


def _normalize_registry_cabinet(payload: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(payload, dict):
        return None
    result = {
        "shkaf_naim": _clean_registry_scalar(payload.get("shkaf_naim"), max_len=200),
        "shkaf_kod": _clean_registry_scalar(payload.get("shkaf_kod"), max_len=80),
        "shkaf_zav_nomer": _clean_registry_scalar(payload.get("shkaf_zav_nomer"), max_len=80),
        "pozicii": [],
    }
    positions = payload.get("pozicii") or []
    if isinstance(positions, list):
        for item in positions[:1000]:
            if not isinstance(item, dict):
                continue
            result["pozicii"].append({
                "nomer": int(item.get("nomer")) if str(item.get("nomer", "")).isdigit() else len(result["pozicii"]) + 1,
                "naimenovanie": _clean_registry_scalar(item.get("naimenovanie"), max_len=220),
                "zavodskoy_nomer": _clean_registry_scalar(item.get("zavodskoy_nomer"), max_len=80),
                "oboznachenie_dok": _clean_registry_scalar(item.get("oboznachenie_dok"), max_len=80),
            })
    if not result["shkaf_naim"] and not result["pozicii"]:
        return None
    return result


def _normalize_registry_evidence(payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    out: Dict[str, Any] = {}
    for field, val in payload.items():
        key = _clean_registry_scalar(field, max_len=80)
        if not key:
            continue
        if key == "zavodskie_nomera" and isinstance(val, list):
            serial_rows = []
            for item in val[:200]:
                if not isinstance(item, dict):
                    continue
                row = {
                    "value": _clean_registry_scalar(item.get("value"), max_len=120),
                }
                page_raw = item.get("page")
                if isinstance(page_raw, int) and page_raw > 0:
                    row["page"] = page_raw
                snippet = _clean_registry_scalar(item.get("snippet"), max_len=240)
                if snippet:
                    row["snippet"] = snippet
                if row["value"]:
                    serial_rows.append(row)
            if serial_rows:
                out[key] = serial_rows
            continue
        if isinstance(val, dict):
            row = {"value": _clean_registry_scalar(val.get("value"), max_len=220)}
            page_raw = val.get("page")
            if isinstance(page_raw, int) and page_raw > 0:
                row["page"] = page_raw
            snippet = _clean_registry_scalar(val.get("snippet"), max_len=320)
            if snippet:
                row["snippet"] = snippet
            if row["value"]:
                out[key] = row
    return out


def _normalize_registry_field_risks(payload: Any) -> Dict[str, Dict[str, str]]:
    if not isinstance(payload, dict):
        return {}
    out: Dict[str, Dict[str, str]] = {}
    for field, val in payload.items():
        if not isinstance(val, dict):
            continue
        key = _clean_registry_scalar(field, max_len=80)
        if not key:
            continue
        level = str(val.get("level") or "").strip().lower()
        if level not in {"low", "medium", "high"}:
            continue
        reason = _clean_registry_scalar(val.get("reason"), max_len=220)
        out[key] = {"level": level, "reason": reason}
    return out


def _normalize_registry_record(raw: Any, idx: int) -> Optional[Dict[str, Any]]:
    if not isinstance(raw, dict):
        return None

    rec: Dict[str, Any] = {}
    for key in REGISTRY_SCALAR_FIELDS:
        rec[key] = _clean_registry_scalar(raw.get(key))
    for key in REGISTRY_LIST_FIELDS:
        val = raw.get(key)
        if not isinstance(val, list):
            val = [val] if val is not None else []
        rec[key] = [_clean_registry_scalar(x, max_len=120) for x in val if _clean_registry_scalar(x, max_len=120)]

    allowed_doc_types = {"single_passport", "group_passport", "cabinet_list", "unknown"}
    rec["document_type"] = rec["document_type"].lower() if rec["document_type"] else "unknown"
    if rec["document_type"] not in allowed_doc_types:
        rec["document_type"] = "unknown"

    rec["zavodskie_nomera"] = normalize_serials(rec.get("zavodskie_nomera") or [])
    rec["normativnye_dok"] = list(dict.fromkeys(rec.get("normativnye_dok") or []))[:40]
    rec["komplektnost"] = list(dict.fromkeys(rec.get("komplektnost") or []))[:40]

    if rec["document_type"] == "group_passport":
        rec["tip_pasporta"] = "group"
    elif rec["zavodskie_nomera"]:
        rec["tip_pasporta"] = "individual"
    else:
        rec["tip_pasporta"] = "no_serial"

    if not rec.get("_fileName"):
        rec["_fileName"] = f"record_{idx}"

    for key in REGISTRY_SCALAR_FIELDS:
        if key in {"document_type", "tip_pasporta", "_fileName"}:
            continue
        if not rec.get(key):
            rec[key] = None

    rec["_saved"] = bool(raw.get("_saved"))
    rec["_edited"] = bool(raw.get("_edited"))

    if isinstance(raw.get("barcode_b64"), str) and 0 < len(raw["barcode_b64"]) <= REGISTRY_B64_MAX:
        rec["barcode_b64"] = raw["barcode_b64"]
    if isinstance(raw.get("barcode_value"), str):
        rec["barcode_value"] = _clean_registry_scalar(raw.get("barcode_value"), max_len=80)

    images: List[str] = []
    raw_images = raw.get("_images")
    if isinstance(raw_images, list):
        for img in raw_images[:REGISTRY_MAX_IMAGES]:
            if isinstance(img, str) and img:
                images.append(img)
    if not images and isinstance(raw.get("_image"), str) and raw["_image"]:
        images = [raw["_image"]]
    if images:
        rec["_images"] = images
        rec["_image"] = images[0]

    view_zoom_raw = raw.get("_viewZoom")
    view_rotation_raw = raw.get("_viewRotation")
    view_page_raw = raw.get("_viewPage")
    view_rotations_raw = raw.get("_viewRotations")
    if isinstance(view_zoom_raw, (int, float)):
        rec["_viewZoom"] = max(0.4, min(4.0, float(view_zoom_raw)))
    if isinstance(view_rotation_raw, (int, float)):
        rec["_viewRotation"] = float(view_rotation_raw)
    if isinstance(view_page_raw, int) and view_page_raw >= 0:
        rec["_viewPage"] = view_page_raw
    if isinstance(view_rotations_raw, list):
        rec["_viewRotations"] = [
            float(v) for v in view_rotations_raw[:REGISTRY_MAX_IMAGES]
            if isinstance(v, (int, float))
        ]

    evidence_map = _normalize_registry_evidence(raw.get("_evidence"))
    if evidence_map:
        rec["_evidence"] = evidence_map

    quality_score, missing_fields, needs_review = _quality_assessment(rec)
    rec["quality_score"] = quality_score
    rec["missing_fields"] = missing_fields
    rec["needs_review"] = needs_review
    rec["_field_risks"] = _build_field_risks(rec, evidence=rec.get("_evidence"))
    raw_field_risks = _normalize_registry_field_risks(raw.get("_field_risks"))
    if raw_field_risks:
        for field, item in raw_field_risks.items():
            current = rec["_field_risks"].get(field)
            rank = {"low": 1, "medium": 2, "high": 3}
            if not current or rank[item["level"]] > rank.get(current.get("level", "low"), 1):
                rec["_field_risks"][field] = item
    rec["_checklist"] = _build_record_checklist(rec)
    return rec


def _registry_empty_state() -> Dict[str, Any]:
    return {
        "version": 1,
        "mode": "p",
        "active_index": None,
        "records": [],
        "cabinet": None,
        "updated_at": _utc_now_iso(),
    }


def _normalize_registry_state(payload: Any) -> Dict[str, Any]:
    base = payload.get("state") if isinstance(payload, dict) and isinstance(payload.get("state"), dict) else payload
    if not isinstance(base, dict):
        return _registry_empty_state()

    records: List[Dict[str, Any]] = []
    raw_records = base.get("records") or []
    if isinstance(raw_records, list):
        for idx, item in enumerate(raw_records[:MAX_REGISTRY_RECORDS], start=1):
            normalized = _normalize_registry_record(item, idx)
            if normalized:
                records.append(normalized)

    cabinet = _normalize_registry_cabinet(base.get("cabinet"))
    mode = "c" if str(base.get("mode", "p")).lower() == "c" else "p"
    active_index_raw = base.get("active_index")
    active_index = active_index_raw if isinstance(active_index_raw, int) else None
    if active_index is not None and (active_index < 0 or active_index >= len(records)):
        active_index = None

    return {
        "version": 1,
        "mode": mode,
        "active_index": active_index,
        "records": records,
        "cabinet": cabinet,
        "updated_at": _utc_now_iso(),
    }


def _load_registry_state() -> Dict[str, Any]:
    path = _registry_path()
    if not path.exists():
        return _registry_empty_state()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning("Failed to read registry state: %s", e)
        return _registry_empty_state()
    return _normalize_registry_state(payload)


def _save_registry_state(payload: Any) -> Dict[str, Any]:
    state = _normalize_registry_state(payload)
    path = _registry_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, path)
    return state


EVAL_DEFAULT_FIELDS = [
    "document_type",
    "naimenovanie",
    "kod_dokumenta",
    "kod_zakaza",
    "data_vypuska",
    "data_priemki",
    "proizvoditel",
    "adres",
    "kontakty",
    "garantia",
    "srok_sluzhby",
    "sertifikat",
    "zavodskie_nomera",
]


def _resolve_control_sample_path(root: Path, safe_name: str) -> Optional[Path]:
    candidate_paths = [
        root / safe_name,
        root / "приложения" / safe_name,
    ]
    source_path = next((p for p in candidate_paths if p.exists()), None)
    if source_path:
        return source_path
    for p in root.rglob(safe_name):
        parts_lower = {part.lower() for part in p.parts}
        if ".venv" in parts_lower or ".git" in parts_lower:
            continue
        return p
    return None


def _make_eval_cache_key(source_path: Path, safe_name: str, allow_llm: Optional[bool], scan_release_date: bool) -> str:
    try:
        st = source_path.stat()
        mt = int(st.st_mtime_ns)
        size = int(st.st_size)
    except Exception:
        mt = 0
        size = 0
    llm_tag = "default" if allow_llm is None else ("on" if allow_llm else "off")
    return (
        f"{source_path.resolve()}::{safe_name}::{mt}::{size}"
        f"::llm={llm_tag}::release={int(scan_release_date)}"
        f"::ocr={OCR_LANG}::psm={','.join(OCR_PSMS)}"
    )


def _eval_cache_get(cache_key: str) -> Optional[Dict[str, Any]]:
    with _eval_extract_cache_lock:
        row = _eval_extract_cache.get(cache_key)
        if not row:
            return None
        payload = row.get("payload")
        if not isinstance(payload, dict):
            return None
        return copy.deepcopy(payload)


def _eval_cache_set(cache_key: str, payload: Dict[str, Any]) -> None:
    with _eval_extract_cache_lock:
        _eval_extract_cache[cache_key] = {
            "stored_at": time.time(),
            "payload": copy.deepcopy(payload),
        }


def _predict_for_control_sample(
    source_path: Path,
    safe_name: str,
    allow_llm: Optional[bool],
    scan_release_date: bool,
    use_cache: bool,
) -> Tuple[Dict[str, Any], bool]:
    cache_key = _make_eval_cache_key(source_path, safe_name, allow_llm=allow_llm, scan_release_date=scan_release_date)
    if use_cache:
        cached = _eval_cache_get(cache_key)
        if cached is not None:
            return cached, True

    with source_path.open("rb") as fh:
        predicted = _extract_document_payload(
            fh.read(),
            safe_name,
            allow_llm=allow_llm,
            scan_release_date=scan_release_date,
        )
    if use_cache:
        _eval_cache_set(cache_key, predicted)
    return predicted, False


def _evaluate_samples(
    samples: List[Dict[str, Any]],
    fields: List[str],
    allow_llm: Optional[bool] = None,
    scan_release_date: bool = True,
    use_cache: bool = False,
) -> Dict[str, Any]:
    total_checks = 0
    matched_checks = 0
    per_field = {f: {"total": 0, "matched": 0} for f in fields}
    sample_reports = []
    cache_hits = 0
    cache_misses = 0

    root = _project_root()
    for idx, sample in enumerate(samples, start=1):
        if not isinstance(sample, dict):
            sample_reports.append({"index": idx, "error": "sample must be object"})
            continue

        filename = str(sample.get("filename") or "").strip()
        expected = sample.get("expected") or {}
        if not filename:
            sample_reports.append({"index": idx, "error": "filename is required"})
            continue
        if not isinstance(expected, dict):
            sample_reports.append({"index": idx, "filename": filename, "error": "expected must be object"})
            continue

        safe_name = os.path.basename(filename)
        if safe_name != filename:
            sample_reports.append({"index": idx, "filename": filename, "error": "unsafe filename"})
            continue

        source_path = _resolve_control_sample_path(root, safe_name)
        if not source_path:
            sample_reports.append({"index": idx, "filename": filename, "error": "file not found"})
            continue

        try:
            predicted, from_cache = _predict_for_control_sample(
                source_path,
                safe_name,
                allow_llm=allow_llm,
                scan_release_date=scan_release_date,
                use_cache=use_cache,
            )
            if from_cache:
                cache_hits += 1
            else:
                cache_misses += 1
        except Exception as e:
            sample_reports.append({"index": idx, "filename": filename, "error": f"extract failed: {e}"})
            continue

        mismatches = []
        checks_for_sample = 0
        matched_for_sample = 0
        for field in fields:
            if field not in expected:
                continue
            exp_val = expected.get(field)
            if exp_val is None or (isinstance(exp_val, str) and not exp_val.strip()):
                continue

            pred_val = predicted.get(field)
            ok = _compare_eval_value(exp_val, pred_val)
            total_checks += 1
            checks_for_sample += 1
            per_field[field]["total"] += 1
            if ok:
                matched_checks += 1
                matched_for_sample += 1
                per_field[field]["matched"] += 1
            else:
                mismatches.append({
                    "field": field,
                    "expected": exp_val,
                    "predicted": pred_val,
                })

        sample_reports.append({
            "index": idx,
            "filename": filename,
            "checks": checks_for_sample,
            "matched": matched_for_sample,
            "accuracy_pct": round((matched_for_sample / checks_for_sample * 100.0), 2) if checks_for_sample else None,
            "mismatches": mismatches,
            "meta": predicted.get("_meta", {}),
            "cached": bool(from_cache),
        })

    per_field_result = {}
    for f, stats in per_field.items():
        total_f = stats["total"]
        matched_f = stats["matched"]
        per_field_result[f] = {
            "total": total_f,
            "matched": matched_f,
            "accuracy_pct": round((matched_f / total_f * 100.0), 2) if total_f else None,
        }

    accuracy_pct = round((matched_checks / total_checks * 100.0), 2) if total_checks else 0.0
    error_rate_pct = round((100.0 - accuracy_pct), 2) if total_checks else 0.0
    return {
        "total_samples": len(samples),
        "total_checks": total_checks,
        "matched_checks": matched_checks,
        "accuracy_pct": accuracy_pct,
        "error_rate_pct": error_rate_pct,
        "per_field": per_field_result,
        "samples": sample_reports,
        "cache": {
            "enabled": bool(use_cache),
            "hits": cache_hits,
            "misses": cache_misses,
        },
        "mode": {
            "allow_llm": ENABLE_LLM if allow_llm is None else bool(allow_llm),
            "scan_release_date": bool(scan_release_date),
        },
    }


@app.route("/api/registry/load", methods=["GET"])
def registry_load():
    return jsonify(_load_registry_state())


@app.route("/api/registry/save", methods=["POST"])
def registry_save():
    payload = request.json or {}
    try:
        state = _save_registry_state(payload)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to save registry: {e}"}), 500
    return jsonify({
        "ok": True,
        "updated_at": state.get("updated_at"),
        "records_count": len(state.get("records") or []),
        "state": state,
    })


@app.route("/api/registry/clear", methods=["POST"])
def registry_clear():
    path = _registry_path()
    try:
        if path.exists():
            path.unlink()
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to clear registry: {e}"}), 500
    return jsonify({
        "ok": True,
        "state": _registry_empty_state(),
    })


@app.route("/api/feedback", methods=["POST"])
def submit_feedback():
    payload = request.json or {}
    message = _clean_registry_scalar(payload.get("message"), max_len=3000)
    topic = _clean_registry_scalar(payload.get("topic"), max_len=120)
    contact = _clean_registry_scalar(payload.get("contact"), max_len=220)
    if not message:
        return jsonify({"error": "Message is required"}), 400

    context = payload.get("context")
    if isinstance(context, dict):
        safe_context = {str(k)[:60]: _clean_registry_scalar(v, max_len=220) for k, v in context.items()}
    else:
        safe_context = {}

    row = {
        "created_at": _utc_now_iso(),
        "topic": topic,
        "contact": contact,
        "message": message,
        "context": safe_context,
    }
    try:
        path = _feedback_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to save feedback: {e}"}), 500
    return jsonify({"ok": True, "created_at": row["created_at"]})


@app.route("/api/evaluate/control", methods=["POST"])
def evaluate_control():
    payload = request.json or {}
    samples = payload.get("samples") or []
    if not isinstance(samples, list) or not samples:
        return jsonify({"error": "Provide non-empty 'samples' list"}), 400

    fields = payload.get("fields") or EVAL_DEFAULT_FIELDS
    if not isinstance(fields, list) or not all(isinstance(x, str) for x in fields):
        return jsonify({"error": "'fields' must be a list of strings"}), 400
    fast = _as_bool(payload.get("fast"), default=EVAL_FAST_DEFAULT)
    use_cache = _as_bool(payload.get("cache"), default=EVAL_USE_CACHE_DEFAULT)
    started = time.perf_counter()
    report = _evaluate_samples(
        samples,
        fields,
        allow_llm=(False if fast else None),
        scan_release_date=not fast,
        use_cache=use_cache,
    )
    report["elapsed_sec"] = round(time.perf_counter() - started, 2)
    report["run_mode"] = "fast" if fast else "full"
    return jsonify(report)


@app.route("/api/evaluate/default", methods=["GET"])
def evaluate_default():
    root = _project_root()
    configured = Path(CONTROL_SAMPLES_FILE)
    candidates = [configured if configured.is_absolute() else (root / configured)]
    candidates.extend([
        root / "samples" / "control_samples.json",
        root / "control_samples.json",
        root / "control_samples.example.json",
    ])
    source = next((p for p in candidates if p.exists()), None)
    if not source:
        return jsonify({"error": "Control samples file not found", "searched": [str(p) for p in candidates]}), 404

    try:
        payload = json.loads(source.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": f"Failed to parse control samples: {e}", "path": str(source)}), 400

    samples = payload.get("samples") if isinstance(payload, dict) else payload
    fields = payload.get("fields") if isinstance(payload, dict) else EVAL_DEFAULT_FIELDS
    if not isinstance(samples, list) or not samples:
        return jsonify({"error": "Control samples must be a non-empty list", "path": str(source)}), 400
    if not isinstance(fields, list) or not all(isinstance(x, str) for x in fields):
        fields = EVAL_DEFAULT_FIELDS

    fast = _as_bool(request.args.get("fast"), default=EVAL_FAST_DEFAULT)
    use_cache = _as_bool(request.args.get("cache"), default=EVAL_USE_CACHE_DEFAULT)
    started = time.perf_counter()
    report = _evaluate_samples(
        samples,
        fields,
        allow_llm=(False if fast else None),
        scan_release_date=not fast,
        use_cache=use_cache,
    )
    report["elapsed_sec"] = round(time.perf_counter() - started, 2)
    report["run_mode"] = "fast" if fast else "full"
    report["source_file"] = str(source)
    return jsonify(report)


def _prepare_1c_rows(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_idx, record in enumerate(records, start=1):
        r = record if isinstance(record, dict) else {}
        serials = normalize_serials(r.get("zavodskie_nomera")) or [""]
        base_key = str(r.get("kod_dokumenta") or r.get("_fileName") or f"record_{row_idx}")
        document_type = str(r.get("document_type") or "").strip().lower()
        quality_raw = r.get("quality_score")
        quality = int(quality_raw) if isinstance(quality_raw, (int, float)) else None

        for sn_idx, serial in enumerate(serials, start=1):
            row = {
                "external_key": f"{base_key}#{sn_idx}",
                "file_name": str(r.get("_fileName") or ""),
                "doc_code": str(r.get("kod_dokumenta") or ""),
                "name": str(r.get("naimenovanie") or ""),
                "order_code": str(r.get("kod_zakaza") or ""),
                "serial_number": str(serial or ""),
                "release_date": str(r.get("data_vypuska") or ""),
                "accept_date": str(r.get("data_priemki") or ""),
                "manufacturer": str(r.get("proizvoditel") or ""),
                "address": str(r.get("adres") or ""),
                "contacts": str(r.get("kontakty") or ""),
                "warranty": str(r.get("garantia") or ""),
                "service_life": str(r.get("srok_sluzhby") or ""),
                "certificate": str(r.get("sertifikat") or ""),
                "document_type": document_type,
                "passport_type": str(r.get("tip_pasporta") or ""),
                "needs_review": bool(r.get("needs_review")),
                "quality_score": quality,
            }
            rows.append(row)

            row_errors: List[str] = []
            if not row["doc_code"]:
                row_errors.append("Не указан код документа")
            if not row["name"]:
                row_errors.append("Не указано наименование")
            if document_type in {"single_passport", "group_passport"} and not row["manufacturer"]:
                row_errors.append("Не указан производитель")
            if document_type == "single_passport" and not row["serial_number"]:
                row_errors.append("Для single_passport отсутствует серийный номер")
            if quality is not None and quality < 60:
                row_errors.append("Низкое качество извлечения (<60)")

            if row_errors:
                errors.append({
                    "row": len(rows),
                    "external_key": row["external_key"],
                    "file_name": row["file_name"],
                    "errors": row_errors,
                })

    return rows, errors


@app.route("/api/export/1c_json", methods=["POST"])
def export_1c_json():
    payload = request.json or {}
    records = payload.get("records", [])
    rows, validation_errors = _prepare_1c_rows(records)
    return jsonify({
        "generated_at": _utc_now_iso(),
        "rows_count": len(rows),
        "errors_count": len(validation_errors),
        "rows": rows,
        "validation_errors": validation_errors,
    })


@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    payload = request.json or {}
    records = payload.get("records", [])
    cabinet = payload.get("cabinet") or {}

    def _safe_join(values: Any) -> str:
        if not values:
            return ""
        if isinstance(values, list):
            return ", ".join(str(v).strip() for v in values if str(v).strip())
        return str(values).strip()

    def _autosize(ws, max_width: int = 68):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Паспорта"
    ws.append([
        "Файл",
        "Тип документа",
        "Тип паспорта",
        "Качество (%)",
        "Нужна проверка",
        "Пробелы в данных",
        "Наименование",
        "Код документа",
        "Код заказа",
        "Дата выпуска",
        "Дата приемки ОТК",
        "Серийные номера",
        "Количество серийных",
        "Производитель",
        "Адрес",
        "Контакты",
        "Гарантия",
        "Срок службы",
        "Сертификат",
        "Нормативные документы",
        "Сохранено",
    ])
    for r in records:
        serials = normalize_serials(r.get("zavodskie_nomera"))
        ws.append([
            r.get("_fileName"),
            r.get("document_type"),
            r.get("tip_pasporta"),
            r.get("quality_score"),
            "Да" if r.get("needs_review") else "Нет",
            _safe_join(r.get("missing_fields")),
            r.get("naimenovanie"),
            r.get("kod_dokumenta"),
            r.get("kod_zakaza"),
            r.get("data_vypuska"),
            r.get("data_priemki"),
            _safe_join(serials),
            len(serials),
            r.get("proizvoditel"),
            r.get("adres"),
            r.get("kontakty"),
            r.get("garantia"),
            r.get("srok_sluzhby"),
            r.get("sertifikat"),
            _safe_join(r.get("normativnye_dok")),
            "Да" if r.get("_saved") else "Нет",
        ])
    ws.freeze_panes = "A2"
    _autosize(ws)

    ws_sn = wb.create_sheet("Серийные номера")
    ws_sn.append(["Файл", "Наименование", "Код документа", "Серийный номер"])
    for r in records:
        serials = normalize_serials(r.get("zavodskie_nomera"))
        for sn in serials:
            ws_sn.append([r.get("_fileName"), r.get("naimenovanie"), r.get("kod_dokumenta"), sn])
    ws_sn.freeze_panes = "A2"
    _autosize(ws_sn)

    ws_norm = wb.create_sheet("Нормативы")
    ws_norm.append(["Файл", "Наименование", "Код документа", "Нормативный документ"])
    for r in records:
        docs = r.get("normativnye_dok") or []
        for doc_name in docs:
            ws_norm.append([r.get("_fileName"), r.get("naimenovanie"), r.get("kod_dokumenta"), doc_name])
    ws_norm.freeze_panes = "A2"
    _autosize(ws_norm)

    ws_1c = wb.create_sheet("1C_Импорт")
    ws_1c.append([
        "ВнешнийКлюч",
        "Файл",
        "КодДокумента",
        "Наименование",
        "КодЗаказа",
        "СерийныйНомер",
        "ДатаВыпуска",
        "ДатаПриемкиОТК",
        "Производитель",
        "Адрес",
        "Контакты",
        "Гарантия",
        "СрокСлужбы",
        "Сертификат",
        "ТипДокумента",
        "ТипПаспорта",
        "НужнаПроверка",
        "Качество",
    ])
    one_c_rows, one_c_errors = _prepare_1c_rows(records)
    for row in one_c_rows:
        ws_1c.append([
            row["external_key"],
            row["file_name"],
            row["doc_code"],
            row["name"],
            row["order_code"],
            row["serial_number"],
            row["release_date"],
            row["accept_date"],
            row["manufacturer"],
            row["address"],
            row["contacts"],
            row["warranty"],
            row["service_life"],
            row["certificate"],
            row["document_type"],
            row["passport_type"],
            "Да" if row["needs_review"] else "Нет",
            row["quality_score"],
        ])
    ws_1c.freeze_panes = "A2"
    _autosize(ws_1c)

    if one_c_errors:
        ws_1c_err = wb.create_sheet("1C_Ошибки")
        ws_1c_err.append(["Строка", "ВнешнийКлюч", "Файл", "Ошибки"])
        for item in one_c_errors:
            ws_1c_err.append([
                item.get("row"),
                item.get("external_key"),
                item.get("file_name"),
                "; ".join(item.get("errors") or []),
            ])
        ws_1c_err.freeze_panes = "A2"
        _autosize(ws_1c_err)

    ws_check = wb.create_sheet("Чек-лист")
    ws_check.append(["Файл", "Критерий", "Статус", "Детали"])
    for r in records:
        items = r.get("_checklist") if isinstance(r.get("_checklist"), list) else _build_record_checklist(r)
        for item in items:
            ws_check.append([
                r.get("_fileName"),
                item.get("title"),
                "OK" if item.get("ok") else "MISS",
                item.get("details"),
            ])
    ws_check.freeze_panes = "A2"
    _autosize(ws_check)

    ws_ev = wb.create_sheet("Источники")
    ws_ev.append(["Файл", "Поле", "Значение", "Страница", "Фрагмент"])
    for r in records:
        ev_map = r.get("_evidence") if isinstance(r.get("_evidence"), dict) else {}
        for field, ev in ev_map.items():
            if field == "zavodskie_nomera" and isinstance(ev, list):
                for item in ev:
                    ws_ev.append([
                        r.get("_fileName"),
                        field,
                        item.get("value"),
                        item.get("page"),
                        item.get("snippet"),
                    ])
            elif isinstance(ev, dict):
                ws_ev.append([
                    r.get("_fileName"),
                    field,
                    ev.get("value"),
                    ev.get("page"),
                    ev.get("snippet"),
                ])
    ws_ev.freeze_panes = "A2"
    _autosize(ws_ev, max_width=90)

    if cabinet:
        ws_cab = wb.create_sheet("Шкаф")
        ws_cab.append([
            "Шкаф",
            "Код шкафа",
            "Заводской номер шкафа",
            "№",
            "Позиция",
            "Заводской номер",
            "Обозначение документа",
        ])
        for item in cabinet.get("pozicii") or []:
            ws_cab.append([
                cabinet.get("shkaf_naim"),
                cabinet.get("shkaf_kod"),
                cabinet.get("shkaf_zav_nomer"),
                item.get("nomer"),
                item.get("naimenovanie"),
                item.get("zavodskoy_nomer"),
                item.get("oboznachenie_dok"),
            ])
        ws_cab.freeze_panes = "A2"
        _autosize(ws_cab)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name="Registry.xlsx", as_attachment=True)


if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    host = os.getenv("FLASK_HOST", "0.0.0.0").strip() or "0.0.0.0"
    try:
        port = int(os.getenv("FLASK_PORT", "5000"))
    except ValueError:
        port = 5000
    app.run(debug=debug, host=host, port=port)


